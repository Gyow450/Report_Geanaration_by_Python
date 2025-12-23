"""
    用来生成PE管定检报告
    默认的文件名为：“PE管定检报告模版.docx”
    
"""
from fastprogress import progress_bar
import openpyxl
from openpyxl.workbook import Workbook
import datetime
import win32com.client as win32
import os
import traceback
import math
from mypackage import r_generator as rg
from mypackage.LOG_DATA import LOG_DICT
from mypackage import interraction_terminal 


"""=========================编辑生成全部用于替换的列表索引文件replacements======================"""


def expand_all_tables(workbook:Workbook, doc, report_name:str)->None:
    """按照读取到的分项报告数量，复制报告页张数。返回穿跨越的组织数量列表"""
    #   楼栋检查报告
    sheet = workbook['结论']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    rows=rg.get_rows_in_sheet(report_name,sheet,log_dict['序号'])
    v=sheet[log_dict['改造地址']+rows[0]].value
    if v=='' or v is None:
        rg.copy_and_insert_report(doc , '复制楼栋检查', 2)
        # rg.copy_and_insert_report(doc , '复制宏观检查报告', 2)
        # rg.copy_and_insert_report(doc , '复制泄漏评估报告', 2)

    rg.replace_text(doc, '复制楼栋检查','',2)
    rg.replace_text(doc, '复制宏观检查报告','',2)
    rg.replace_text(doc, '复制泄漏评估报告','',2)

    
    #   整理删除页面
    # rg.delete_page_by_text(doc, '待删除')


def expand_all_figs(workbook:Workbook, doc, report_name:str):
    """根据‘全部静态台账工作簿’中的内容，复制空白图张数"""
    sheet = workbook['管道基本信息']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    times:int = sheet[log_dict['路由图数量']+rows[0]].value
    # times:int = 39
    if times>1:
        rg.copy_and_insert_report(doc , '+复制路由图', times)
    rg.replace_text(doc, '+复制路由图','',2)

def do_add_row_for_all(report_name:str,doc,workbook:Workbook):
    """执行‘管道清单’，‘问题清单’两个表格的扩张和填入"""
    if config['是否写入管道清单']:
        sheet = workbook['管道清单']
        log_dict:dict = rg.get_col_in_sheet(sheet)
        rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号']) 
        rg.add_row_to_table(doc,'管道序号',len(rows)-1)
        print('填写管道清单')
        rg.write_in_table(doc,'管道序号',sheet,rows)

    print('填写问题清单')
    # 遍历宏观检查中的问题
    sheet = workbook['宏观检查记录']
    white_list:set[str] = {'无','完好','符合','正常','合格'}
    log_dict:dict = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号']) 
    all_problems:list[str] = list()
    for row in rows:
        son_rows:list[str] = rg.get_rows_in_sheet(sheet[log_dict['记录自编号']+row].value,sheet,log_dict['所属记录编号'])
        for son_row in son_rows:
            for key_word in ['地面标志','管道防护带','地表环境','阀门井']:
                problem:str|None = sheet[log_dict[key_word]+son_row].value
                if problem is not None and problem not in white_list:
                    v = [
                        sheet[log_dict['管段（桩号）']+row].value.strip(),
                        sheet[log_dict['地表参照及位置描述']+son_row].value.strip(),
                        f"{sheet[log_dict['坐标X']+son_row].value} ,{sheet[log_dict['坐标Y']+son_row].value}",
                        key_word+'：'+sheet[log_dict[key_word]+son_row].value,
                        ]
                    all_problems.append(v)
    depth_list:list[tuple[str]]=[]
    
    for row in rows:
        son_rows:list[str] = rg.get_rows_in_sheet(sheet[log_dict['记录自编号']+row].value,sheet,log_dict['所属记录编号'])
        for son_row in son_rows:
            depth = sheet[log_dict['管道埋深']+son_row].value
            if depth is not None and depth !='':
                if sheet[log_dict['埋深达标']+son_row].value == '埋深不足':
                    depth_list += [(
                                    sheet[log_dict['地表参照及位置描述']+son_row].value.strip(),
                                    f"{float(sheet[log_dict['坐标X']+son_row].value):.0f} ，{float(sheet[log_dict['坐标Y']+son_row].value):.0f}",
                                    f"{float(depth):.1f}",
                                    '埋深不足'
                                    )] 
                else:
                    depth_list += [(
                                    sheet[log_dict['地表参照及位置描述']+son_row].value.strip(),
                                    f"{float(sheet[log_dict['坐标X']+son_row].value):.0f} ，{float(sheet[log_dict['坐标Y']+son_row].value):.0f}",
                                    f"{float(depth):.1f}",
                                    '符合'
                                    )] 
          
    #   遍历开挖检测中的问题
    sheet = workbook['开挖检测记录']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号']) 
    # for row in rows:
    #     text:str = ''
    #     temp_dict:dict[str,str] = dict()
    #     for key_word in ['缺陷描述','热熔连接缺陷描述','电熔连接缺陷描述']:
    #         temp_dict[key_word] = sheet[log_dict[key_word]+row].value
    #     for key_word,word in temp_dict.items():
    #         if word is not None:
    #             text += key_word+'——'+word+'\n'
    #     if text !='':       
    #         v:list[str] = [
    #             sheet[log_dict['探坑位置']+row].value,
    #             str(sheet[log_dict['探坑坐标 X']+row].value) +','+str(sheet[log_dict['探坑坐标 Y']+row].value),
    #             text
    #             ]
    #         all_problems.append(v)
    for row in rows:
        depth = sheet[log_dict['管道埋深（m）']+row].value
        
        if sheet[log_dict['埋深达标']+row].value == '埋深不足':
            depth_list += [(
                            sheet[log_dict['探坑位置']+row].value.strip(),
                            f"{float(sheet[log_dict['探坑坐标 X']+row].value):.0f} ，{float(sheet[log_dict['探坑坐标 Y']+row].value):.0f}",
                            f"{float(depth):.1f}",
                            '埋深不足'
                            )] 
        else:
            depth_list += [(
                            sheet[log_dict['探坑位置']+row].value.strip(),
                            f"{float(sheet[log_dict['探坑坐标 X']+row].value):.0f} ，{float(sheet[log_dict['探坑坐标 Y']+row].value):.0f}",
                            f"{float(depth):.1f}",
                            '符合'
                            )] 
    #   执行写入所有问题
    rg.add_row_to_table(doc,'问题序号',len(all_problems)-1)
    for table in doc.Tables:
        first_cell = table.Cell(1,1)
        fist_cell_text:str|None = first_cell.Range.Text.strip()
        if '问题序号' in fist_cell_text:
            i:int=1
            for problem in all_problems:
                i += 1
                table.Cell(i,1).Range.Text = i-1
                j:int = 2
                for problem_text in problem:
                    if problem_text == None :
                        problem_text='不明'
                    table.Cell(i,j).Range.Text = problem_text
                    j += 1
            break

    rg.add_row_to_table(doc,'埋深序号',len(depth_list)-1)          
    for table in doc.Tables:
        first_cell = table.Cell(1,1)
        fist_cell_text:str|None = first_cell.Range.Text.strip()
        if '埋深序号' in fist_cell_text:
            i = 1
            for s_depth in depth_list:
                i += 1
                table.Cell(i,1).Range.Text = i-1
                j:int = 2
                for depth_text in s_depth:
                    if depth_text == None :
                        depth_text='不明'
                    table.Cell(i,j).Range.Text = depth_text
                    j += 1
            break

"""
========================执行替换=======================
"""
#   替换文本
def do_replace(doc , replacements1:list[tuple[str,str]],replacements2:list[tuple[str,str]]=[])->None:
    """替换所有文本，先替换全局，再替换单次"""
    for target_text, replacement_text in replacements2:
        rg.replace_text(doc, target_text, replacement_text,2 )
        if target_text in '+报告编码':
            for section in doc.Sections:
        # 遍历节中的所有页眉
                for header in section.Headers:
                    # 替换页眉中的文本
                    if "+报告编码" in header.Range.Text:
                        header.Range.Find.ClearFormatting()
                        header.Range.Find.Replacement.ClearFormatting()
                        header.Range.Find.Execute(
                            FindText="+报告编码",
                            MatchCase=False,
                            MatchWholeWord=True,
                            MatchWildcards=False,
                            MatchSoundsLike=False,
                            MatchAllWordForms=False,
                            Forward=True,
                            Wrap=1,  # wdFindStop
                            Format=False,
                            ReplaceWith=replacement_text,
                            Replace=2  # wdReplaceAll
                        )
    for target_text, replacement_text in replacements1:
        rg.replace_text(doc, target_text, replacement_text )

#   签字函数
def make_sign_dig_log(workbook:Workbook,doc,report_name:str,path:str,cross_list:list[int])->dict:
    """编制签名图片、开挖图片、路由图替换的索引"""
    # 签结论页、资料审查：遍历所有的宏观和开挖记录检验人员，通过集合无序化，结论页最多签4人，资料最多签2人
    sign_dict:dict = {}
    all_names:set[str] = set()
    sheet = workbook['宏观检查记录']
    dict_log:dict[str,str] = rg.get_col_in_sheet(sheet)
    rows:list[str] =rg.get_rows_in_sheet(report_name,sheet,dict_log['报告编号'])
    for row in rows:
        if sheet[dict_log['检验人员']+row].value:
            names:list[str] =sheet[dict_log['检验人员']+row].value.split(',')
        for name in names:
            all_names.add(name)

    sheet = workbook['开挖检测记录']
    dict_log:dict[str,str] = rg.get_col_in_sheet(sheet)
    rows:list[str] =rg.get_rows_in_sheet(report_name,sheet,dict_log['报告编号'])
    for row in rows:
        if sheet[dict_log['检验人员']+row].value:
            names:list[str] =sheet[dict_log['检验人员']+row].value.split(',')
        else:
            names=[]
        for name in names:
            all_names.add(name)
    sign_dict['签字'] = []  
    temp_list:list[str] =list(all_names)
    lenth = len(temp_list)
    sign_dict['签字'] += temp_list[:4]  # 结论页检验人员
    if lenth<4:
        sign_dict['签字'] += ['空白']*(4-lenth) 
    sign_dict['签字'] += temp_list[:1]  # 编制人员
    sign_dict['签字'] += temp_list[:2]  # 资料审查
    if lenth<2:
        sign_dict['签字'] += ['空白']*(2-lenth)

    # 签宏观检查报告
    sheet =workbook['宏观检查记录']
    dict_log = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,dict_log['报告编号'])
    for row in rows:
        names_set:set[str] = set()
        if  sheet[dict_log['检验人员']+row].value:
            for name in sheet[dict_log['检验人员']+row].value.split(', '):
                names_set.add(name)
        lenth = len(names_set)
        temp_list = list(names_set)
        sign_dict['签字'] += temp_list[:2]  # 宏观检查
        if lenth<2:
            sign_dict['签字'] += ['空白']*(2-lenth)

    #   签开挖检验报告
    sheet =workbook['开挖检测记录']
    dict_log = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,dict_log['报告编号'])
    sign_dict['开挖']=[] 
    for row in rows:
        sign_dict['开挖'].append(sheet[dict_log['记录自编号']+row].value) 
        names_set = set()
        if sheet[dict_log['检验人员']+row].value:
            for name in sheet[dict_log['检验人员']+row].value.split(','):
                names_set.add(name)
        lenth = len(names_set)
        temp_list = list(names_set)
        sign_dict['签字'] += temp_list[:2]  
        if lenth<2:
            sign_dict['签字'] += ['空白']*(2-lenth)


    #   穿跨越记录
    sheet =workbook['宏观检查记录']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    i:int = 0
    if all(x==0 for x in cross_list):
        sign_dict['签字']+=list(all_names)[:2]
    for row in rows:
        names_set:set = set()
        if sheet[log_dict['检验人员']+row].value:
            for name in sheet[log_dict['检验人员']+row].value.split(', '):
                names_set.add(name)
        for _ in range(cross_list[i]):
            lenth = len(names_set)
            temp_list = list(names_set)
            sign_dict['签字'] += temp_list[:2]  
            if lenth<2:
                sign_dict['签字'] += ['空白']*(2-lenth)
        i += 1
    
    #   风险评估报告
    any_name=list(all_names)[0]
    sign_dict['签字'] += [any_name]*2
   
    #   替换路由图：根据报告编号来定位
    sign_dict['管道分图']=[]
    sign_dict['管道总图']=[]
    if config['是否写入管道路由图']:
        sheet = workbook['管道基本信息']
        log_dict = rg.get_col_in_sheet(sheet)
        rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
        street_name = sheet[log_dict['街道']+rows[0]].value
        count = sheet[log_dict['路由图纸张数']+rows[0]].value
        image_path = path+'\\800公里报告图片\\管网分公司_'+street_name+'.jpeg'
        # image_path = path+'\\新繁\\管道路由图\\新繁总图.jpg'
        # image_path = path+'\\大丰\\大丰2025定检图纸\\大丰公司.jpg'
        sign_dict['管道总图'].append(image_path)
        for i in range(count): 
            image_path = path+'\\800公里报告图片\\管网分公司_'+street_name+'-'+str(i+1)+'.jpeg' 
        # for i in range(44):
        #     image_path = path+'\\大丰\\大丰2025定检图纸\\大丰公司_DF'+str(i+1)+'.jpeg'  
        # for i in range(39):
        #     image_path = path+'\\新繁\\管道路由图\\新繁_XF'+str(i+1)+'.jpeg'   
            sign_dict['管道分图'].append(image_path)
    return sign_dict

#   替换报告图片
# def do_replace_picture(workbook:Workbook,doc,report_name:str,path:str)->None:
#     """实际执行替换开挖报告中的照片"""
#     sheet = workbook['开挖检测记录']
#     log_dict:dict[str,str] = rg.get_col_in_sheet(sheet)
#     rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
#     for row in rows:
#         image_name:str = f"{path}\\管网840\\开挖照片\\{sheet[log_dict['记录自编号']+row].value}"
#         # image_name:str = f"{path}\\新繁\\开挖照片\\{sheet[log_dict['记录自编号']+row].value}"
#         pic_ex_names:list[str] = ['.jpg','.png','.jpeg']
#         for ex_name in pic_ex_names:
#             image_path = f"{image_name}{ex_name}"
#             if os.path.exists(image_path):
#                 rg.insert_picture(doc,image_path,'开挖检测',120)
#                 break

#   替换管道路由图
# def  do_replace_figs(workbook:Workbook,doc,report_name:str,path:str)->None:
#     """替换路由图：根据报告编号来定位"""
#     sheet = workbook['全部静态台账']
#     log_dict = rg.get_col_in_sheet(sheet)
#     rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
#     street_name = sheet[log_dict['街道']+rows[0]].value
#     count = sheet[log_dict['路由图纸张数']+rows[0]].value
#     image_path = path+'\\800公里报告图片\\管网分公司_'+street_name+'.jpeg'
#     rg.insert_picture(doc,image_path,'管道总图',580)
#     for i in range(count): 
#         image_path = path+'\\800公里报告图片\\管网分公司_'+street_name+'-'+str(i+1)+'.jpeg'  
#         rg.insert_picture(doc,image_path,'管道分图',580)
    
    #     rg.insert_picture(doc,image_path,'管道分图',580)

    #     rg.insert_picture(doc,image_path,'管道分图',580)
    

#   完成索引
def make_replacement_index(workbook:Workbook,report_name:str)->dict:
    """全体替换内容，主要调整函数"""
    replacements:dict={}
    replacements['文本'] = []
    #   正文
    temp_list:list[tuple]= []
    sheet =workbook['结论']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['编号'])
    row = rows[0]
    global_name:str = f"{sheet[log_dict['评估单元']+row].value}立管"
    case_code:int = sheet[log_dict['case']+row].value
    temp_text:str = f"{global_name}评估项目共包含{sheet[log_dict['总户数']+row].value}户。"
    replacements['文本'] += [('+项目概况',temp_text)]
    temp_text:str = f"通过资料审查发现：{global_name}部分管道设计、施工资料缺失"
    if sheet[log_dict['年限']+row].value >=30:
        temp_text+='，实际运行年限较长'
    replacements['文本'] += [('+写入审查问题',f"{temp_text}。")]
    # temp_loop_text:str=''
    # for key in ['4级及以上锈蚀','立管被完全封闭','敷设位置和方式不满足规范要求']:
    #     value=sheet[log_dict[key]+row].value
    #     if value >0:
    #         temp_loop_text += f"{value}户存在{key}，"
    # temp_loop_text=rg.check_text(temp_loop_text)
    # temp_text:str = f"通过宏观检查发现：{global_name}立管共{temp_loop_text}"
    temp_text:str = ''
    if case_code == 0:
        temp_text+='部分管道腐蚀泄漏严重，连接处防腐较差。'
    elif case_code==1:
        temp_text+=f"{sheet[log_dict['改造地址']+row].value}部分管道腐蚀泄漏严重，连接处防腐较差。"
    elif case_code==2:
        temp_text+=f"{sheet[log_dict['改造地址']+row].value}部分管道连接处防腐有破损。"
    replacements['文本'] += [('+写入宏观检查',temp_text)]
    temp_text:str = f"通过泄漏发现：{global_name}立管无浓度反应。"
    replacements['文本'] += [('+写入泄漏评估',temp_text)]
    problem_text:str=''
    temp_text:str=f"{global_name}立管建议{sheet[log_dict['结论']+row].value}。"
    replacements['文本'] += [('+评估结论',temp_text)]
   
    #   明设评估表
    replacements['老化评估-明设评估报告']=[]
    temp_list:list[tuple[str,str|datetime.datetime]]=[]


    #   评估结论表
    replacements['老化评估结论表']=[]
    temp_list:list[tuple[str,str|datetime.datetime]]=[]
    temp_list += [('对象简述',f"{global_name}评估项目共包含{sheet[log_dict['总户数']+row].value}。")]
    temp_list += [('长度m',f"{sheet[log_dict['总户数']+row].value*3}m")]
    temp_list += [('管材类别',f"{sheet[log_dict['管道材质']+row].value}")]
    temp_list += [('使用单位',f"{sheet[log_dict['评估单元']+row].value}用户")]
    if case_code<1:
        temp_list += [('$限期改造','☑限期改造')]
        temp_list += [('$落实安全管控措施，可继续运行','□落实安全管控措施，可继续运行')]  
    else:
        temp_list += [('$限期改造','□限期改造')]
        temp_list += [('$落实安全管控措施，可继续运行','☑落实安全管控措施，可继续运行')]    
    temp_list += [('$立即改造','□立即改造')]
    temp_list += [('$符合安全运行要求','□符合安全运行要求')]
    temp_list += [('$材质落后','□材质落后')] 
    temp_list += [('$使用年限较长','☑使用年限较长')]
    problem_text+='使用年限较长，处于/临近人员密集区域，防腐状况较差，'
    temp_list += [('$防腐状况较差','☑防腐状况较差')]
    temp_list += [('$建构筑物占压','□建构筑物占压')]
    temp_list += [('$处于/临近地质灾害易发区域','□处于/临近地质灾害易发区域')]
    temp_list += [('$处于/临近人员密集区','☑处于/临近人员密集区')]
    temp_list += [('$其他主要问题：','□其他主要问题：/')]
    if case_code<2:
        temp_list += [('$腐蚀泄漏严重','☑腐蚀泄漏严重')]
        problem_text += '腐蚀泄露严重'
    else:
        temp_list += [('$腐蚀泄漏严重','□腐蚀泄漏严重')]
    replacements['老化评估结论表'].append(temp_list)
    
    #   补充正文的结论
    temp_text:str=f"通过对以上单项检测评估结果进行综合评定，{global_name}主要存在的问题为：{rg.check_text(problem_text)}。"
    replacements['文本'] += [('+写入评估主要问题',temp_text)]
    
    #   资料审查报告 
    replacements['老化评估-资料审查报告']=[]
    temp_list:list[tuple]= []
    for key in LOG_DICT['老化评估-资料审查报告']:
        if '+' not in key:
            if key in log_dict:
                temp_list +=[(key,sheet[log_dict[key]+row].value)]
            else:
                temp_list +=[(key,'不明')]
    replacements['老化评估-资料审查报告'].append(temp_list)

    #   宏观检查报告
    replacements['老化评估-宏观检查报告']=[]
    temp_list:list[tuple]= []
    for key in LOG_DICT['老化评估-宏观检查报告']:
        if '+' not in key:
            if key in log_dict:
                temp_list +=[(key,sheet[log_dict[key]+row].value)]
            else:
                temp_list +=[(key,'不明')]
    if case_code==0:
        temp_list += [('+结论',f"结论及发现问题：接口处管道防腐层部分损坏，多处管道本体腐蚀严重。")]
    elif case_code==1:
        temp_list += [('+结论',f"结论及发现问题：接口处管道防腐层部分损坏，{sheet[log_dict['改造地址']+row].value}管道本体腐蚀。")]
    else:
        temp_list += [('+结论',f"结论及发现问题：接口处管道防腐层部分损坏，管道本体轻微腐蚀。")]
        
    replacements['老化评估-宏观检查报告'].append(temp_list)

    #   泄漏评估报告
    sheet =workbook['资料']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['编号'])
    replacements['老化评估-泄漏评估报告']=[]    
    temp_list:list[tuple]= []
    for key in LOG_DICT['老化评估-泄漏评估报告']:
        if '+' not in key:
            if key in log_dict:
                temp_list +=[(key,sheet[log_dict[key]+row].value)]
            else:
                temp_list +=[(key,'不明')]
    replacements['老化评估-泄漏评估报告'].append(temp_list)
    return replacements

def make_all_replacement_index(workbook,report_name):
    """管道基本信息：报告编号、管道名称、管道长度等"""
    replacements:list = []
    sheet = workbook['资料']
    log_dict:dict =rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['编号'])
    row = rows[0]
    # # lenth:int = sheet[log_dict['大于20年长度']+row].value + sheet[log_dict['小于20年长度']+row].value
    # lenth:int = sheet[log_dict['管道长度']+row].value 
    # l:float = lenth/1000
    replacements += [
                ('+报告编码',f"PG-{report_name}"),
                ('+管道类型','立管'),
                ('+使用单位','重庆江津天然气有限责任公司'),
                ('+管道名称',sheet[log_dict['管道名称']+row].value),
                ('+评估日期','2022年11月')
                # ('+使用单位',sheet[log_dict['使用单位']+row].value),
                # # ('+使用单位','成都燃气集团股份有限公司管网分公司'),
                # ('+检验日期',sheet[log_dict['检验日期']+row].value),
                # ('+管道名称',sheet[log_dict['管道名称']+row].value),
                # ('+管道长度',l),
                # ('+管道长度','164.12'),# 新繁
                # ('+管道名称','天然气管道'),
                # ('+检验日期','2025年06月15日') # 新繁
                # ('+使用单位','成都成燃新繁燃气有限公司'),
                 ]
    # 检查是否有不明管段
    # sheet = workbook['管段清单']
    # log_dict:dict =rg.get_col_in_sheet(sheet)
    # rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    # count_unkown:int =0
    # for row in rows:
    #     if '使用单位指定管段' in sheet[log_dict['工程名称']+row].value:
    #         count_unkown =1
    #         break
    # if count_unkown==1:
    #     replacements+=[('+不明管道检查','部分管段无资料，其余管段仅见竣工图')]
    # else:
    #     replacements+=[('+不明管道检查','所有管段仅见竣工图')]
    # replacements+=[('+不明管道检查','所有管段仅见竣工图')]
    # used_years:list[float] = []
    # for row in rows:
    #     if sheet[log_dict['实际使用年限']+row].value:
    #         used_years += [sheet[log_dict['实际使用年限']+row].value]
    # replacements+=[('+投运年限',f"{min(used_years)}—{max(used_years)}年")]
    return replacements

def do_replace_in_son_report(doc,any_dict):
    """执行分项报告表格写入"""
    i:int = 0
    j:int = 0
    k:int = 0
    l:int = 0
    for table in doc.Tables:
        title_name:str = table.Title
        if title_name == '老化评估结论表':
            rg.replace_text_in_table(doc,table,any_dict['老化评估结论表'][i],'老化评估结论表')  
            i+=1
        elif title_name == '老化评估-资料审查报告':
            rg.replace_text_in_table(doc,table,any_dict['老化评估-资料审查报告'][j],'老化评估-资料审查报告')  
            j+=1
        elif title_name == '老化评估-宏观检查报告':
            rg.replace_text_in_table(doc,table,any_dict['老化评估-宏观检查报告'][k],'老化评估-宏观检查报告')  
            k+=1
        elif title_name == '老化评估-泄漏评估报告':
            rg.replace_text_in_table(doc,table,any_dict['老化评估-泄漏评估报告'][l],'老化评估-泄漏评估报告')  
            l+=1
        

def do_replace_all_pic(doc,pic_dict:dict,path:str):
    """执行所有图片的替换"""
    i:int = 0
    j:int = 0
    k:int = 0
    for shape in doc.InlineShapes:
        tag:str = shape.Title 
        if tag == '签字':
            if config['是否生成签字']:
                if pic_dict['签字'][i]=='空白':
                    pass
                else:
                    rg.replace_pictue(doc,f"{config['签名图片所在']}\\{pic_dict['签字'][i]}.png",shape)
            i+=1
        elif tag == '开挖':
            for ex_name in ['.jpg','.png','.jpeg']:
                # f_path:str = f"{path}\\管网840\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                # f_path:str = f"{path}\\新繁\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                # f_path:str = f"{path}\\大丰\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                # f_path:str = f"{path}\\郫三司\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                f_path:str = f"{config['数据源所在']}\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                if os.path.exists(f_path):
                    rg.replace_pictue(doc,f_path,shape,120)
                    break
            j+=1
        elif tag == '管道总图':
            if config['是否写入管道路由图']:
                rg.replace_pictue(doc,pic_dict['管道总图'][0],shape,580)
            pass
        elif tag == '管道分图':
            if config['是否写入管道路由图']:
                rg.replace_pictue(doc,pic_dict['管道分图'][k],shape,580) 
            k+=1
        else:
            pass
    
def solo_main(report_name:str,workbook:Workbook,word,path:str):

    replacements_dict:dict = {}
    replacements_list:list[tuple] = []
    doc_modle_path = f"{config['模板文件']}"
    try:
        doc = word.Documents.Open(doc_modle_path)
               
        print('生成替换用文本')
        replacements_dict |= make_replacement_index(workbook,report_name)
        replacements_list += make_all_replacement_index(workbook,report_name) 

        print('替换内容')
        do_replace( doc , replacements_dict['文本'],replacements_list )
        
        # print('附件表格处理')
        # do_add_row_for_all(report_name,doc,workbook)

        print('扩张分项报告表格')
        expand_all_tables(workbook, doc, report_name)

        # print('替换残余内容')
        # do_replace( doc , replacements_dict['文本a'])
    
        print('填写分项报告表格')
        do_replace_in_son_report(doc,replacements_dict)
    
        # if config['是否写入管道路由图']:
        #     print('扩张路由图')
        #     expand_all_figs(workbook, doc, report_name)

        # print('编制图片替换索引') 
        # sign_dict=make_sign_dig_log(workbook,doc,report_name,path,cross_list)

        # print('替换所有图片')
        # do_replace_all_pic(doc,sign_dict,path)


        # 移动到文档的末端
        selection = word.Selection
        selection.EndKey(6)  # 6 表示 wdStory，即整个文档

        # 更新文档中的所有域
        doc.Fields.Update()
        
        output_file = f"{config['输出文件所在']}\\{report_name}.docx"
        doc.SaveAs2(output_file, FileFormat=16)  # 16 表示docx 17 表示 PDF
        
        # output_file = f"{config['输出文件所在']}\\{report_name}.pdf"
        # doc.SaveAs2(output_file, FileFormat=17)  
        
        print(f"文档已保存为：{output_file}")

    except Exception as ex:
        traceback.print_exc()
        if doc is not None:
            doc.SaveAs2(f"{config['输出文件所在']}\\error_{report_name}.docx",FileFormat =16)
            print(f"{report_name}发生错误！")
            doc.Saved =True
            raise ex
    finally:
        if doc is not None:
            doc.Close(SaveChanges=False)


if __name__ == '__main__':
    config:dict[str,str|bool]={
        '模板文件':'',
        '数据源文件':'',
        # '数据源所在':'',
        # '签名图片所在':'',
        '输出文件所在':'',
        # '是否生成概述段落':False,
        # '是否写入封面':False,
        # '是否写入管道清单':False,
        # '是否写入管道路由图':False,
        # '是否生成签字':False,    
    }
    set_list:list[tuple[int,str,str|bool]]=[
        (2,'模板文件','docx'),
        (2,'数据源文件','xlsx'),
        # (0,'数据源所在',''),
        # (0,'签名图片所在',''),
        (0,'输出文件所在',''),
        # (3,'是否生成概述段落',False),
        # (3,'是否写入封面',False),
        # (3,'是否写入管道清单',False),
        # (3,'是否写入管道路由图',False),
        # (3,'是否生成签字',False),    
    ]
    config=interraction_terminal.set_argumments(set_list)
    app_type = rg.check_office_installation()
    if app_type == None:
        print('未找到合适的应用以打开文档')

    path:str = os.getcwd()
    
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\犀浦\\犀浦_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\新繁\\新繁_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\大丰\\大丰_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\管网840\\管网840_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\郫三司\\郫三司_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{config['数据源所在']}\\原始数据.xlsx" )
    workbook:Workbook = openpyxl.load_workbook(f"{config['数据源文件']}" )

    print('读取模板文件')
    
    if app_type == "office":
        word = win32.Dispatch("Word.Application")
    elif app_type == "wps":
        word = win32.Dispatch("Kwps.Application")
    word.Visible = False  # 不显示 Word 窗口，加快处理速度
    word.DisplayAlerts = 0  # 关闭警告信息
    # 全局关闭拼写/语法检查
    word.Options.CheckSpellingAsYouType = False   # 关闭实时拼写检查
    word.Options.CheckGrammarAsYouType = False    # 关闭实时语法检查
    word.Options.ContextualSpeller = False        # 关闭上下文拼写检查（Word 2010+）
    #   初始化完成
    sheet=workbook['结论']
    all_names:list[str]=[]
    log_dict =rg.get_col_in_sheet(sheet)
    for cell in sheet[log_dict['编号']]:    # 遍历静态台账里所有编号
        v:str = cell.value
        if v is not None:
            all_names.append(v)
        else:
            break
    
    for report_name in progress_bar( all_names[1:]):
        try:
            solo_main(report_name,workbook,word,path)
        except Exception as e:
            print('有错误发生')
        finally:
            continue
    # report_name = 'DGB2025001CD'
    # solo_main(report_name,workbook,word,path)



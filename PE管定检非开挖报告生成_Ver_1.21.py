"""
    用来生成PE管定检报告
    默认的文件名为：“PE管定检报告模版.docx”
    
"""
import openpyxl
from openpyxl.workbook import Workbook
import datetime
import win32com.client as win32
import os
import traceback
import math
from mypackage import r_generator as rg
from mypackage.LOG_DATA import LOG_DICT,RISKY_EVA_C,RISKY_EVA_S
from mypackage import set_config 


"""=========================编辑生成全部用于替换的列表索引文件replacements======================"""

def make_text_for_c1(report_name:str,workbook:Workbook):
    """用于生成每个工程最开始的概况文本"""
    sheet= workbook['管段清单']
    log_dict:dict[str] = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    c1_list:list[str] = []
    key_cols:list[str] = [
                    log_dict['工程名称'],
                    '，起止点坐标',log_dict['起止点坐标'],
                    '，竣工图号',log_dict['竣工图号'],
                    '，竣工日期',log_dict['竣工日期'],
                    '，运行年限',log_dict['实际使用年限'],
                    '。'
                    ]
    i:int=0
    for row in rows:
        i += 1
        text:str = rg.get_text_by_log(sheet,row,key_cols)
        c1_list.append(f"{i}、{text}")

    return c1_list

def expand_all_tables(workbook:Workbook, doc, report_name:str)->list[int]:
    """按照读取到的分项报告数量，复制报告页张数。返回穿跨越的组织数量列表"""
    #   宏观检查报告
    sheet = workbook['宏观检查记录']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    times:int = len(rg.get_rows_in_sheet(report_name, sheet ,log_dict['报告编号'])) 
    if times>1:
        rg.copy_and_insert_report(doc , '复制宏观检查报告', times)
    rg.replace_text(doc, '复制宏观检查报告','',2)

    #   开挖检测
    sheet = workbook['开挖检测记录']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    times:int = len(rg.get_rows_in_sheet(report_name , sheet ,log_dict['报告编号']))
    if times>1:
        rg.copy_and_insert_report(doc , '复制开挖报告', times)
    rg.replace_text(doc, '复制开挖报告','',2)  
       
    #   穿跨越检查
    sheet = workbook['宏观检查记录']
    log_dict:dict[str,str] = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    cross_list:list[int]=[] 
    times:int = 0
    for row in rows:
        count:str|None =sheet[log_dict['穿越总结']+row].value
        if count:
            ctrl_value:int =max(math.ceil(count.count('穿越')/9),math.ceil(count.count('跨越')/5))
            times += ctrl_value
            cross_list.append(ctrl_value)
        else:
            cross_list.append(0)
    if times>1:
        rg.copy_and_insert_report(doc , '复制穿跨越报告', times)
    rg.replace_text(doc, '复制穿跨越报告','',2)  
    return cross_list
    
    #   整理删除页面
    # rg.delete_page_by_text(doc, '待删除')


def expand_all_figs(workbook:Workbook, doc, report_name:str):
    """根据‘全部静态台账工作簿’中的内容，复制空白图张数"""
    sheet = workbook['管道基本信息']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    times:int = sheet[log_dict['路由图数量']+rows[0]].value
    # times:int = 39
    if times>1:
        rg.copy_and_insert_report(doc , '+复制路由图', times)
    rg.replace_text(doc, '+复制路由图','',2)

def do_add_row_for_all(report_name:str,doc,workbook:Workbook):
    """执行‘管道清单’，‘问题清单’两个表格的扩张和填入"""
    if config['list']:
        sheet = workbook['管道清单']
        log_dict:dict = rg.get_col_in_sheet(sheet)
        rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号']) 
        rg.add_row_to_table(doc,'管道序号',len(rows)-1)
        print('填写管道清单')
        rg.write_in_table(doc,'管道序号',sheet,rows)

    print('填写问题清单')
    # 遍历宏观检查中的问题
    sheet = workbook['宏观检查记录']
    white_list:set[str] = {'无','完好','符合','正常','合格'}
    log_dict:dict = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号']) 
    all_problems:list[str] = list()
    for row in rows:
        son_rows:list[str] = rg.get_rows_in_sheet(sheet[log_dict['记录自编号']+row].value,sheet,log_dict['所属记录编号'])
        for son_row in son_rows:
            for key_word in ['地面标志','管道防护带','地表环境','阀门井']:
                problem:str|None = sheet[log_dict[key_word]+son_row].value
                if problem is not None and problem not in white_list:
                    v = [
                        sheet[log_dict['管段（桩号）']+row].value.strip(),
                        sheet[log_dict['地表参照及位置描述']+son_row].value.strip(),
                        f"{sheet[log_dict['坐标X']+son_row].value} ,{sheet[log_dict['坐标Y']+son_row].value}",
                        key_word+'：'+sheet[log_dict[key_word]+son_row].value,
                        ]
                    all_problems.append(v)
    depth_list:list[tuple[str]]=[]
    
    for row in rows:
        son_rows:list[str] = rg.get_rows_in_sheet(sheet[log_dict['记录自编号']+row].value,sheet,log_dict['所属记录编号'])
        for son_row in son_rows:
            depth = sheet[log_dict['管道埋深']+son_row].value
            if depth is not None and depth !='':
                if sheet[log_dict['埋深达标']+son_row].value == '埋深不足':
                    depth_list += [(
                                    sheet[log_dict['地表参照及位置描述']+son_row].value.strip(),
                                    f"{float(sheet[log_dict['坐标X']+son_row].value):.0f} ，{float(sheet[log_dict['坐标Y']+son_row].value):.0f}",
                                    f"{float(depth):.1f}",
                                    '埋深不足'
                                    )] 
                else:
                    depth_list += [(
                                    sheet[log_dict['地表参照及位置描述']+son_row].value.strip(),
                                    f"{float(sheet[log_dict['坐标X']+son_row].value):.0f} ，{float(sheet[log_dict['坐标Y']+son_row].value):.0f}",
                                    f"{float(depth):.1f}",
                                    '符合'
                                    )] 
          
    #   遍历开挖检测中的问题
    sheet = workbook['开挖检测记录']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号']) 
    # for row in rows:
    #     text:str = ''
    #     temp_dict:dict[str,str] = dict()
    #     for key_word in ['缺陷描述','热熔连接缺陷描述','电熔连接缺陷描述']:
    #         temp_dict[key_word] = sheet[log_dict[key_word]+row].value
    #     for key_word,word in temp_dict.items():
    #         if word is not None:
    #             text += key_word+'——'+word+'\n'
    #     if text !='':       
    #         v:list[str] = [
    #             sheet[log_dict['探坑位置']+row].value,
    #             str(sheet[log_dict['探坑坐标 X']+row].value) +','+str(sheet[log_dict['探坑坐标 Y']+row].value),
    #             text
    #             ]
    #         all_problems.append(v)
    for row in rows:
        depth = sheet[log_dict['管道埋深（m）']+row].value
        
        if sheet[log_dict['埋深达标']+row].value == '埋深不足':
            depth_list += [(
                            sheet[log_dict['探坑位置']+row].value.strip(),
                            f"{float(sheet[log_dict['探坑坐标 X']+row].value):.0f} ，{float(sheet[log_dict['探坑坐标 Y']+row].value):.0f}",
                            f"{float(depth):.1f}",
                            '埋深不足'
                            )] 
        else:
            depth_list += [(
                            sheet[log_dict['探坑位置']+row].value.strip(),
                            f"{float(sheet[log_dict['探坑坐标 X']+row].value):.0f} ，{float(sheet[log_dict['探坑坐标 Y']+row].value):.0f}",
                            f"{float(depth):.1f}",
                            '符合'
                            )] 
    #   执行写入所有问题
    rg.add_row_to_table(doc,'问题序号',len(all_problems)-1)
    for table in doc.Tables:
        first_cell = table.Cell(1,1)
        fist_cell_text:str|None = first_cell.Range.Text.strip()
        if '问题序号' in fist_cell_text:
            i:int=1
            for problem in all_problems:
                i += 1
                table.Cell(i,1).Range.Text = i-1
                j:int = 2
                for problem_text in problem:
                    if problem_text == None :
                        problem_text='不明'
                    table.Cell(i,j).Range.Text = problem_text
                    j += 1
            break

    rg.add_row_to_table(doc,'埋深序号',len(depth_list)-1)          
    for table in doc.Tables:
        first_cell = table.Cell(1,1)
        fist_cell_text:str|None = first_cell.Range.Text.strip()
        if '埋深序号' in fist_cell_text:
            i = 1
            for s_depth in depth_list:
                i += 1
                table.Cell(i,1).Range.Text = i-1
                j:int = 2
                for depth_text in s_depth:
                    if depth_text == None :
                        depth_text='不明'
                    table.Cell(i,j).Range.Text = depth_text
                    j += 1
            break

"""
========================执行替换=======================
"""
#   替换文本
def do_replace(doc , replacements1:list[tuple[str,str]],replacements2:list[tuple[str,str]]=[])->None:
    """替换所有文本，先替换全局，再替换单次"""
    for target_text, replacement_text in replacements2:
        rg.replace_text(doc, target_text, replacement_text,2 )
    for target_text, replacement_text in replacements1:
        rg.replace_text(doc, target_text, replacement_text )

#   签字函数
def make_sign_dig_log(workbook:Workbook,doc,report_name:str,path:str,cross_list:list[int])->dict:
    """编制签名图片、开挖图片、路由图替换的索引"""
    # 签结论页、资料审查：遍历所有的宏观和开挖记录检验人员，通过集合无序化，结论页最多签4人，资料最多签2人
    sign_dict:dict = {}
    all_names:set[str] = set()
    sheet = workbook['宏观检查记录']
    dict_log:dict[str,str] = rg.get_col_in_sheet(sheet)
    rows:list[str] =rg.get_rows_in_sheet(report_name,sheet,dict_log['报告编号'])
    for row in rows:
        if sheet[dict_log['检验人员']+row].value:
            names:list[str] =sheet[dict_log['检验人员']+row].value.split(',')
        for name in names:
            all_names.add(name)

    sheet = workbook['开挖检测记录']
    dict_log:dict[str,str] = rg.get_col_in_sheet(sheet)
    rows:list[str] =rg.get_rows_in_sheet(report_name,sheet,dict_log['报告编号'])
    for row in rows:
        if sheet[dict_log['检验人员']+row].value:
            names:list[str] =sheet[dict_log['检验人员']+row].value.split(',')
        else:
            names=[]
        for name in names:
            all_names.add(name)
    sign_dict['签字'] = []  
    temp_list:list[str] =list(all_names)
    lenth = len(temp_list)
    sign_dict['签字'] += temp_list[:4]  # 结论页检验人员
    if lenth<4:
        sign_dict['签字'] += ['空白']*(4-lenth) 
    sign_dict['签字'] += temp_list[:1]  # 编制人员
    sign_dict['签字'] += temp_list[:2]  # 资料审查
    if lenth<2:
        sign_dict['签字'] += ['空白']*(2-lenth)

    # 签宏观检查报告
    sheet =workbook['宏观检查记录']
    dict_log = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,dict_log['报告编号'])
    for row in rows:
        names_set:set[str] = set()
        if  sheet[dict_log['检验人员']+row].value:
            for name in sheet[dict_log['检验人员']+row].value.split(', '):
                names_set.add(name)
        lenth = len(names_set)
        temp_list = list(names_set)
        sign_dict['签字'] += temp_list[:2]  # 宏观检查
        if lenth<2:
            sign_dict['签字'] += ['空白']*(2-lenth)

    #   签开挖检验报告
    sheet =workbook['开挖检测记录']
    dict_log = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,dict_log['报告编号'])
    sign_dict['开挖']=[] 
    for row in rows:
        sign_dict['开挖'].append(sheet[dict_log['记录自编号']+row].value) 
        names_set = set()
        if sheet[dict_log['检验人员']+row].value:
            for name in sheet[dict_log['检验人员']+row].value.split(','):
                names_set.add(name)
        lenth = len(names_set)
        temp_list = list(names_set)
        sign_dict['签字'] += temp_list[:2]  
        if lenth<2:
            sign_dict['签字'] += ['空白']*(2-lenth)


    #   穿跨越记录
    sheet =workbook['宏观检查记录']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    i:int = 0
    if all(x==0 for x in cross_list):
        sign_dict['签字']+=list(all_names)[:2]
    for row in rows:
        names_set:set = set()
        if sheet[log_dict['检验人员']+row].value:
            for name in sheet[log_dict['检验人员']+row].value.split(', '):
                names_set.add(name)
        for _ in range(cross_list[i]):
            lenth = len(names_set)
            temp_list = list(names_set)
            sign_dict['签字'] += temp_list[:2]  
            if lenth<2:
                sign_dict['签字'] += ['空白']*(2-lenth)
        i += 1
    
    #   风险评估报告
    any_name=list(all_names)[0]
    sign_dict['签字'] += [any_name]*2
   
    #   替换路由图：根据报告编号来定位
    sign_dict['管道分图']=[]
    sign_dict['管道总图']=[]
    if config['pip_fig']:
        sheet = workbook['管道基本信息']
        log_dict = rg.get_col_in_sheet(sheet)
        rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
        street_name = sheet[log_dict['街道']+rows[0]].value
        count = sheet[log_dict['路由图纸张数']+rows[0]].value
        image_path = path+'\\800公里报告图片\\管网分公司_'+street_name+'.jpeg'
        # image_path = path+'\\新繁\\管道路由图\\新繁总图.jpg'
        # image_path = path+'\\大丰\\大丰2025定检图纸\\大丰公司.jpg'
        sign_dict['管道总图'].append(image_path)
        for i in range(count): 
            image_path = path+'\\800公里报告图片\\管网分公司_'+street_name+'-'+str(i+1)+'.jpeg' 
        # for i in range(44):
        #     image_path = path+'\\大丰\\大丰2025定检图纸\\大丰公司_DF'+str(i+1)+'.jpeg'  
        # for i in range(39):
        #     image_path = path+'\\新繁\\管道路由图\\新繁_XF'+str(i+1)+'.jpeg'   
            sign_dict['管道分图'].append(image_path)
    return sign_dict

#   替换报告图片
# def do_replace_picture(workbook:Workbook,doc,report_name:str,path:str)->None:
#     """实际执行替换开挖报告中的照片"""
#     sheet = workbook['开挖检测记录']
#     log_dict:dict[str,str] = rg.get_col_in_sheet(sheet)
#     rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
#     for row in rows:
#         image_name:str = f"{path}\\管网840\\开挖照片\\{sheet[log_dict['记录自编号']+row].value}"
#         # image_name:str = f"{path}\\新繁\\开挖照片\\{sheet[log_dict['记录自编号']+row].value}"
#         pic_ex_names:list[str] = ['.jpg','.png','.jpeg']
#         for ex_name in pic_ex_names:
#             image_path = f"{image_name}{ex_name}"
#             if os.path.exists(image_path):
#                 rg.insert_picture(doc,image_path,'开挖检测',120)
#                 break

#   替换管道路由图
# def  do_replace_figs(workbook:Workbook,doc,report_name:str,path:str)->None:
#     """替换路由图：根据报告编号来定位"""
#     sheet = workbook['全部静态台账']
#     log_dict = rg.get_col_in_sheet(sheet)
#     rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
#     street_name = sheet[log_dict['街道']+rows[0]].value
#     count = sheet[log_dict['路由图纸张数']+rows[0]].value
#     image_path = path+'\\800公里报告图片\\管网分公司_'+street_name+'.jpeg'
#     rg.insert_picture(doc,image_path,'管道总图',580)
#     for i in range(count): 
#         image_path = path+'\\800公里报告图片\\管网分公司_'+street_name+'-'+str(i+1)+'.jpeg'  
#         rg.insert_picture(doc,image_path,'管道分图',580)
    
    #     rg.insert_picture(doc,image_path,'管道分图',580)

    #     rg.insert_picture(doc,image_path,'管道分图',580)
    

#   完成索引
def make_replacement_index(workbook:Workbook,report_name:str)->dict:
    """全体替换内容，主要调整函数"""
    replacements:dict={}
    replacements['文本'] = []
    #   封面
    temp_list:list[tuple]= []
    sheet =workbook['管道基本信息']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    row = rows[0]
    for key,value in LOG_DICT['PE定期检验报告']:
        text = sheet[log_dict[value]+row].value
        temp_list.append((key,text))
    replacements['文本'] += temp_list
    global_name = sheet[log_dict['管道名称']+row].value
    #   结论页，数据仍出自基本信息
    temp_list= []
    temp_list = [
                ('+使用单位',sheet[log_dict['使用单位']+row].value),
                ('+单位地址',sheet[log_dict['单位地址']+row].value),
                ('+安全管理人员',sheet[log_dict['安全管理人员']+row].value),
                ('+联系电话',sheet[log_dict['联系电话']+row].value),
                ('+编制日期',sheet[log_dict['检验日期']+row].value),
                ]
    replacements['文本'] += temp_list
    
    #   正文部分
    if config['list']:
        sheet = workbook['管段清单']
        log_dict = rg.get_col_in_sheet(sheet)
        rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
        replacements['文本'] += [('+工程总数',len(rows))]*3
   
    #   统计测深点数，埋深不足点数，是否有深根植物、土壤扰动、管道占压，是否有穿跨越保护，遍历日期
    depth_count1=depth_count2=0
    error_depth1=error_depth2=0
    sheet =workbook['宏观检查记录']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号']) 
    all_date:list=list()
    check_other:dict= {'深根植物伴行':0,'土壤扰动':0,'管道占压':0}
    any_pass:int =0    # 穿跨越保护状态
    any_pass_ex:int=0   # 穿跨越存在性
    for row in rows:
        v_text = sheet[log_dict['穿、跨越公路总结']+row].value
        ex_v_text = sheet[log_dict['穿越总结']+row].value
        if ex_v_text =='' or ex_v_text is None:
            pass
        else:
            any_pass_ex = 1
        if v_text is None:
            pass
        elif '保护设施完好' in v_text:
            any_pass=1
        v_text = sheet[log_dict['穿、跨越河流总结']+row].value
        if v_text is None:
            pass
        elif '保护设施完好' in v_text:
            any_pass=1
        v_text = sheet[log_dict['管道防护带总结']+row].value
        if v_text is None:
            pass
        elif '深根植物' in v_text:
            check_other['深根植物伴行']=1
        v_text = sheet[log_dict['地表环境总结']+row].value
        if v_text is None:
            pass
        elif '土壤扰动' in v_text:
            check_other['土壤扰动']=1
        elif '管道占压' in v_text:
            check_other['管道占压']=1
        v=sheet[log_dict['检验日期']+row].value
        if v!='' and v is not None:
            all_date.append(v)
        record_num:str = sheet[log_dict['记录自编号']+row].value
        son_rows:list[str] = rg.get_rows_in_sheet(record_num,sheet,log_dict['所属记录编号'])
        for son_row in son_rows:
            v = sheet[log_dict['管道埋深']+son_row].value
            u = sheet[log_dict['埋深达标']+son_row].value
            if v is not None:
                depth_count1 += 1
            if u is not None:
                error_depth1 += 1
    another_text:str=''
    for key,value in check_other.items():
        if value ==0:
            pass
        else:
            another_text+=f"{key}、"
    if another_text=='':
        replacements['文本'].append(('+环境检查结果','以上检查内容未见异常'))
    else:
        replacements['文本'].append(('+环境检查结果',f"部分管段存在{rg.check_text(another_text)}等情况，其余检查内容未见异常。"))
    if any_pass_ex == 0:
        replacements['文本'].append(('+穿、跨越检查结果','本次检验管道无穿、跨越段'))
    elif any_pass==0:
        replacements['文本'].append(('+穿、跨越检查结果','无穿、跨越管道保护设施'))
    else:
        replacements['文本'].append(('+穿、跨越检查结果','穿、跨越保护设施完好'))
    

    next_ins_year:int = 4
    sheet =workbook['开挖检测记录']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    for row in rows:
         v = sheet[log_dict['检验日期']+row].value
         if v !='' and v is not None: 
            all_date.append(v)           
         u = sheet[log_dict['埋深达标']+row].value
         if u is not None:
                error_depth2 += 1
         w = sheet[log_dict['结论']+row].value
         if '3级' in w or '4级' in w:
             next_ins_year = 3
             
    depth_count2 += len(rows)
    if not all_date:
        all_date = [datetime.datetime.strptime( '2024年10月15日',"%Y年%m月%d日"),]
    first_date = min(all_date)
    last_date = max(all_date)
    next_time = first_date + datetime.timedelta(days=next_ins_year*365-31)  #下次检验日期
    replacements['文本'] += [
                    ('+测深数量',depth_count1+depth_count2),
                    ('+开挖总数',depth_count2),
                    ('+最早日期',first_date),
                    # ('+检验日期',first_date),   #   这里填充的是资料审查报告里的检验日期
                    ('+下次检验时间',next_time.strftime("%Y年%m月")),
                    ('+下次检验时间',next_time.strftime("%Y年%m月")),
                     ]
    error_depth= error_depth1+error_depth2
    if error_depth >0:
        replacements['文本'].append(('埋深均符合要求',str(error_depth)+'处埋深不符合要求'))
   
    
    
    #   宏观检查报告 
    replacements['宏观检查报告']=[]
    replacements['文本a']=[]
    temp_list:list[tuple]= []
    sheet =workbook['宏观检查记录']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])  
    temp_count:int = 0  #分项报告编号
    white_list:set[str] = {'无','符合','完好','合格','正常','保护设施完好'}
    for row in rows:
        
        temp_count += 1
        replacements['文本a']+=[('报告#',f"报告（{temp_count}）")]
        temp_list = [
                    ('+管道名称',global_name),
                    ('+管段',sheet[log_dict['管段（桩号）']+row].value),
                    ('+设备名称型号',sheet[log_dict['设备名称型号']+row].value),
                    ('+设备编号',sheet[log_dict['设备编号']+row].value),
                    # ('+检验日期',sheet[log_dict['检验日期']+row].value),
                    ]
        results = ''
        
        for key in LOG_DICT['宏观检查报告']:        #   具体检查项
            result:str = ''
            options:tuple[str] = LOG_DICT['宏观检查报告'][key]      #   体系里所设定：检查项的各种异常选项
            s_text:str|None = sheet[log_dict[key+'总结']+row].value #   多维表格里的单项总结文本 
            if s_text is None or s_text =='':
                check_list:list = []
            elif ',' in s_text:
                check_list:list[str] = s_text.split(',')     #   单选项的总结，用英文,隔断
            else:
                check_list:list[str] = s_text.split(' ')     #   复选项的总结，用空格隔断
            temp_text:str = ''                          #   承载勾选框的文本
            if key not in sheet[log_dict['检查项目总结']+row].value:    #   报告整体总结中的无此项检查
                if '无此项' in options:                 #   若有无此项可选则选，如无则在白名单中匹配
                    for option in options:
                        if option =='无此项':
                            temp_text += f"☑{option}、"
                        else:
                            temp_text += f"□{option}、"
                else:
                    for option in options:
                        if option in white_list:
                            temp_text += f"☑{option}、"
                        else:
                            temp_text += f"□{option}、"
                temp_list += [(f"+{key}",f"{temp_text}□",'  ')]
            else:
                if set(check_list)-white_list:   #   并非只有白名单值
                    extra_options:set[str] =set(check_list)-set(options)
                    for option in options:
                        if option in check_list and option not in white_list:
                            temp_text += f"☑{option}、"
                        else:
                            temp_text += f"□{option}、"
                    if extra_options and '保护设施完好' not in check_list:                   #   存在额外选项
                        if '全线无标志' in extra_options:
                            temp_text=temp_text.replace('□无标志','☑无标志')
                            temp_list += [(f"+{key}",f"{temp_text}□",'  ')]
                        elif '全线深根植物伴行' in extra_options:
                            temp_text=temp_text.replace('□深根植物','☑深根植物')
                            extra_options.discard('全线深根植物伴行')
                            if extra_options:
                                temp_list += [(f"+{key}",f"{temp_text}☑",'，'.join(extra_options))]
                            else:
                                temp_list += [(f"+{key}",f"{temp_text}□",'  ')]
                        else:
                            temp_list += [(f"+{key}",f"{temp_text}☑",'，'.join(extra_options))]
                    else:
                        temp_list += [(f"+{key}",f"{temp_text}□",'  ')]
                    
                        
                else:                                   #   只有白名单值
                    for option in options:
                        if option in white_list:
                            temp_text += f"☑{option}、"
                        else:
                            temp_text += f"□{option}、"
                    temp_list += [(f"+{key}",f"{temp_text}□",'  ')]
            # 统计问题数量
            if s_text != '' and s_text is not None:
                for prblem in set(check_list):
                    if prblem in white_list or prblem in {'无跨越、穿越段仅路面宏观检验','无跨越、穿越段仅宏观检验','暗渠上方跨越，仅地表宏观检查'}:
                        pass
                    else:
                        if '全线无标志' in prblem or '全线深根植物伴行' in prblem: 
                            r_text = f"{prblem}"
                        else:
                            r_text = f"{check_list.count(prblem)}处{prblem}"
                        result += f"{r_text}，"
            if result !='':
                result=rg.check_text(result)
                results += f"{key}存在问题：{result}；"
        
        if results != '':
            results = f"{rg.check_text(results)}；示踪装置：无示踪线。"
        else:
            # results = '以上项目宏观检查未发现异常。'
            results = '示踪装置：无示踪线。'
        temp_list = temp_list + [('+结论',f"结论：{results}")] 
        replacements['宏观检查报告'].append(temp_list) 

    
    #   开挖检验报告
    replacements['开挖检验报告']=[]
    sheet = workbook['开挖检测记录']
    log_dict = rg.get_col_in_sheet(sheet)  #获取表头索引
    
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    temp_count = 0
    for row in rows:
        temp_list:list = []
        temp_count += 1
        replacements['文本a'] +=[('报告#',f"报告（{temp_count}）")]
        temp_list += rg.make_change_text_for_heading(sheet,row,'开挖检测记录',log_dict)
        temp_list += [('+管道名称',f"{global_name}（{sheet[log_dict['探坑位置']+row].value}）")]
        temp_list += rg.make_change_text_for_option(sheet,row,'开挖检测记录',log_dict)
        v1 = sheet[log_dict['备注']+row].value
        if v1 is None:
            v1 = ''
        else:
            v1 = f"{rg.check_text(v1)}。"
        v2 = rg.check_text(sheet[log_dict['结论']+row].value)
        temp_list += [('+备注',f"备注：{v1}根据GB/T 43922-2024《在役聚乙烯燃气管道检验与评价》安全状况等级评定为{v2}。")]
        replacements['开挖检验报告'].append(temp_list)
    #   穿、跨越检查
    replacements['穿、跨越报告'] = []
    sheet =workbook['宏观检查记录']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    temp_count = 0  # 报告标题的编号
    report_no:int = 0       # 穿跨越报告的序号
    if any_pass_ex == 0:
        replacements['文本a'] +=[('报告#',"报告")]
        temp_list=[('+管段','全部管段'),
                # ('+检验日期',sheet[log_dict['检验日期']+row].value),
                ('+检查结论',"检查结论：本次检验的管道无穿、跨越段"),
                ('&号1','/'),
                ('&长度1','/'),
                ('&发现问题及位置描述1','/'),
                ('&备注1','/'),
                ('$号1','/'),
                ('$长度1','/'),
                ('$发现问题及位置描述1','/'),
                ('$备注1','/')
                ]
        replacements['穿、跨越报告'].append(temp_list)               
        

    for row in rows:
        # 本地确认
        exists_bool:str|None = sheet[log_dict['穿越总结']+row].value
        if exists_bool == None:
            continue
        else:
            report_no+=1
            record_num = sheet[log_dict['记录自编号']+row].value
            son_rows = rg.get_rows_in_sheet(record_num,sheet,log_dict['所属记录编号'])
            son_rows1:list[str]=[]
            son_rows2:list[str]=[]
            for son_row in son_rows:
                v = sheet[log_dict['穿跨越类型']+son_row].value
                if v == '跨越':
                    son_rows1.append(son_row)
                elif v=='穿越':
                    son_rows2.append(son_row)
            cap1:int=len(son_rows1)
            cap2:int=len(son_rows2)
            pages:int = max(math.ceil(cap1/5),math.ceil(cap2/9))
            cr_result1 = sheet[log_dict['穿、跨越河流总结']+row].value
            cr_result2 = sheet[log_dict['穿、跨越公路总结']+row].value
            b_result = '穿、跨越段仅宏观检验,未发现异常'
            if cr_result1 is not None and '保护设施完好' in cr_result1:
                b_result = '保护设施完好'
            if cr_result2 is not None and '保护设施完好' in cr_result2:
                b_result = '保护设施完好'
            page:int = 0        # 标志单一报告的页序数
            for _ in range(pages):
                page += 1
                if pages ==1:
                    replacements['文本a'] +=[('报告#',f"报告（{report_no}）")]
                else:
                    replacements['文本a'] +=[('报告#',f"报告（{report_no}-{page}）")]
                temp_list:list[tuple]=[]
                temp_list += [      # 表头和结论
                        ('+管段',sheet[log_dict['管段（桩号）']+row].value),
                        # ('+检验日期',sheet[log_dict['检验日期']+row].value),
                        ('+检查结论',f"检查结论：{b_result}"),
                        ]
                # 跨越填表
                temp_count = 0
                if len(son_rows1) >0:
                    for son_row in son_rows1[:5]:
                        temp_count +=1
                        temp_list +=[
                            (f"&号{temp_count}",temp_count+5*(page-1)),
                            (f'&长度{temp_count}',sheet[log_dict['穿跨越长度']+son_row].value),
                            (f'&发现问题及位置描述{temp_count}',sheet[log_dict['地表参照及位置描述']+son_row].value ),
                            (f'&备注{temp_count}','/')
                        ]
                    if temp_count < 5:
                        temp_count += 1
                        temp_list += [
                            (f"&号{temp_count}",'/'),
                            (f'&长度{temp_count}','/'),
                            (f'&发现问题及位置描述{temp_count}','/' ),
                            (f'&备注{temp_count}','/')
                        ]
                    son_rows1 = son_rows1[5:]
                else:
                    temp_list += [
                            ("&号1",'/'),
                            ('&长度1','/'),
                            ('&发现问题及位置描述1','/' ),
                            ('&备注1','/')
                        ]
                # 穿越填表
                temp_count = 0
                if len(son_rows2) >0:
                    for son_row in son_rows2[:9]:
                        temp_count +=1
                        temp_list +=[
                            (f"$号{temp_count}",temp_count+9*(page-1)),
                            (f'$长度{temp_count}',sheet[log_dict['穿跨越长度']+son_row].value),
                            (f'$发现问题及位置描述{temp_count}',sheet[log_dict['地表参照及位置描述']+son_row].value ),
                            (f'$备注{temp_count}','/')
                        ]
                    if temp_count < 9:
                        temp_count += 1
                        temp_list += [
                            (f"$号{temp_count}",'/'),
                            (f'$长度{temp_count}','/'),
                            (f'$发现问题及位置描述{temp_count}','/' ),
                            (f'$备注{temp_count}','/')
                        ]
                        son_rows2 = son_rows2[9:]
                else:
                    temp_list += [
                            ("$号1",'/'),
                            ('$长度1','/'),
                            ('$发现问题及位置描述1','/' ),
                            ('$备注1','/')
                        ]
                replacements['穿、跨越报告'].append(temp_list)
    # 风险评估
    sheet =workbook['风险评估']
    log_dict=rg.get_col_in_sheet(sheet)
    rows =rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    replacements['风险预评估']=[]
    row = rows[0]
    for key_str,list_dict in RISKY_EVA_S.items():
        risk_score:int=0 
        for son_key,son_tuple in list_dict.items():
            v= sheet[log_dict[son_key]+row].value # 表格实际内容
            for option,score in son_tuple:
                if isinstance(option,tuple): # 如果键是区间（元组）
                    if v>=option[0] and v<option[1]:
                        risk_score += score
                else:
                    if v == option:
                        risk_score += score
        replacements['风险预评估'].append(risk_score)
    for key_str,list_dict in RISKY_EVA_C.items():
        risk_score:int=0 
        v= sheet[log_dict[key_str]+row].value # 表格实际内容
        for any_tuple in list_dict:
            option,score =any_tuple
            if isinstance(option,tuple): # 如果键是区间（元组）
                if v>=option[0] and v<option[1]:
                    risk_score += score
            else:
                if v == option:
                    risk_score += score 
        replacements['风险预评估'].append(risk_score)
    s_sigma_value = sum(replacements['风险预评估'][:8])
    c_sigma_value = sum(replacements['风险预评估'][8:])
    r_value = s_sigma_value*c_sigma_value
    if r_value<3600:
        r_class='低风险'
    elif r_value>=3600 and r_value<7800:
        r_class='中风险'
    elif r_value>=7800 and r_value<12600:
        r_class='较高风险'
    else:
        r_class='高风险'

    replacements['文本']+=[
        ('+预评估失效可能性得分',s_sigma_value),
        ('+预评估失效后果得分',c_sigma_value),
        ('+预评估风险值',r_value),
        ('+预评估风险等级',r_class)
        ]
    
    replacements['风险再评估']=[]
    row = rows[1]
    for key_str,list_dict in RISKY_EVA_S.items():
        risk_score:int=0 
        for son_key,son_tuple in list_dict.items():
            v= sheet[log_dict[son_key]+row].value # 表格实际内容
            for option,score in son_tuple:
                if isinstance(option,tuple): # 如果键是区间（元组）
                    if v>=option[0] and v<option[1]:
                        risk_score += score
                else:
                    if v == option:
                        risk_score += score
        replacements['风险再评估'].append(risk_score)
    for key_str,list_dict in RISKY_EVA_C.items():
        risk_score:int=0 
        v= sheet[log_dict[key_str]+row].value # 表格实际内容
        for any_tuple in list_dict:
            option,score = any_tuple
            if isinstance(option,tuple): # 如果键是区间（元组）
                if v>=option[0] and v<option[1]:
                    risk_score = score
            else:
                if v == option:
                    risk_score = score 
        replacements['风险再评估'].append(risk_score)
    s_sigma_value = sum(replacements['风险再评估'][:8])
    c_sigma_value = sum(replacements['风险再评估'][8:])
    r_value = s_sigma_value*c_sigma_value
    if r_value<3600:
        r_class='低风险'
    elif r_value>=3600 and r_value<7800:
        r_class='中风险'
    elif r_value>=7800 and r_value<12600:
        r_class='较高风险'
    else:
        r_class='高风险'
    replacements['文本']+=[
        ('+再评估失效可能性得分',s_sigma_value),
        ('+再评估失效后果得分',c_sigma_value),
        ('+再评估风险值',r_value),
        ('+再评估风险等级',r_class),
        ('+再评估风险等级',r_class),
        ('+再评估风险等级',r_class)
        ]
    return replacements

def make_all_replacement_index(workbook,report_name):
    """管道基本信息：报告编号、管道名称、管道长度等"""
    replacements:list = []
    sheet = workbook['管道基本信息']
    log_dict:dict =rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    row = rows[0]
    # lenth:int = sheet[log_dict['大于20年长度']+row].value + sheet[log_dict['小于20年长度']+row].value
    lenth:int = sheet[log_dict['管道长度']+row].value 
    l:float = lenth/1000
    replacements += [
                ('+报告编号',report_name),
                ('+使用单位',sheet[log_dict['使用单位']+row].value),
                # ('+使用单位','成都燃气集团股份有限公司管网分公司'),
                ('+检验日期',sheet[log_dict['检验日期']+row].value),
                ('+管道名称',sheet[log_dict['管道名称']+row].value),
                ('+管道长度',l),
                # ('+管道长度','164.12'),# 新繁
                # ('+管道名称','天然气管道'),
                # ('+检验日期','2025年06月15日') # 新繁
                # ('+使用单位','成都成燃新繁燃气有限公司'),
                 ]
    # 检查是否有不明管段
    # sheet = workbook['管段清单']
    # log_dict:dict =rg.get_col_in_sheet(sheet)
    # rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    # count_unkown:int =0
    # for row in rows:
    #     if '使用单位指定管段' in sheet[log_dict['工程名称']+row].value:
    #         count_unkown =1
    #         break
    # if count_unkown==1:
    #     replacements+=[('+不明管道检查','部分管段无资料，其余管段仅见竣工图')]
    # else:
    #     replacements+=[('+不明管道检查','所有管段仅见竣工图')]
    # replacements+=[('+不明管道检查','所有管段仅见竣工图')]
    # used_years:list[float] = []
    # for row in rows:
    #     if sheet[log_dict['实际使用年限']+row].value:
    #         used_years += [sheet[log_dict['实际使用年限']+row].value]
    # replacements+=[('+投运年限',f"{min(used_years)}—{max(used_years)}年")]
    return replacements

def do_replace_in_son_report(doc,any_dict):
    """执行分项报告表格写入"""
    i:int = 0
    j:int = 0
    k:int = 0
    l:int = 1
    m:int = 1
    for table in doc.Tables:
        title_name:str = table.Title
        if title_name == '宏观检查报告':
            rg.replace_text_in_table(doc,table,any_dict['宏观检查报告'][i],'宏观检查报告索引')  
            i+=1
        elif title_name == '开挖检验报告':
            rg.replace_text_in_table(doc,table,any_dict['开挖检验报告'][j],'开挖检测报告索引')  
            j+=1
        elif title_name == '穿、跨越报告':
            rg.replace_text_in_table(doc,table,any_dict['穿、跨越报告'][k],'穿、跨越报告索引')  
            k+=1
        elif title_name == '风险预评估报告':
            son_table = table.Cell(3,1).Tables(1)
            for score in any_dict['风险预评估'][:8]:
                cell = son_table.Cell(2,l)
                cell.Range.Text = score
                l+=1
            son_table = table.Cell(3,1).Tables(2)
            for score in any_dict['风险预评估'][8:]:
                cell = son_table.Cell(2,l-8)
                cell.Range.Text = score
                l+=1
        elif title_name == '风险再评估报告':
            son_table = table.Cell(3,1).Tables(1)
            for score in any_dict['风险再评估'][:8]:
                cell = son_table.Cell(2,m)
                cell.Range.Text = score
                m+=1
            son_table = table.Cell(3,1).Tables(2)
            for score in any_dict['风险再评估'][8:]:
                cell = son_table.Cell(2,m-8)
                cell.Range.Text = score
                m+=1
        else:
            pass

def do_replace_all_pic(doc,pic_dict:dict,path:str):
    """执行所有图片的替换"""
    i:int = 0
    j:int = 0
    k:int = 0
    for shape in doc.InlineShapes:
        tag:str = shape.Title 
        if tag == '签字':
            if config['sign']:
                if pic_dict['签字'][i]=='空白':
                    pass
                else:
                    rg.replace_pictue(doc,f"{config['sign_path']}\\{pic_dict['签字'][i]}.png",shape)
            i+=1
        elif tag == '开挖':
            for ex_name in ['.jpg','.png','.jpeg']:
                # f_path:str = f"{path}\\管网840\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                # f_path:str = f"{path}\\新繁\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                # f_path:str = f"{path}\\大丰\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                # f_path:str = f"{path}\\郫三司\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                f_path:str = f"{config['dir']}\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                if os.path.exists(f_path):
                    rg.replace_pictue(doc,f_path,shape,120)
                    break
            j+=1
        elif tag == '管道总图':
            if config['pip_fig']:
                rg.replace_pictue(doc,pic_dict['管道总图'][0],shape,580)
            pass
        elif tag == '管道分图':
            if config['pip_fig']:
                rg.replace_pictue(doc,pic_dict['管道分图'][k],shape,580) 
            k+=1
        else:
            pass

def do_input_para(doc,workbook:Workbook,report_name:str)->None:
    sheet = workbook['管段清单']
    log_dict = rg.get_col_in_sheet(sheet)
    rg.copy_and_insert_paragraph(doc,'复制写入概况',make_text_for_c1(report_name,workbook))
    
    sheet =workbook['宏观检查记录']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    check1:int =0   # 遗失阀井
    check2:int =0   # 深根植物 
    for row in rows:
        if sheet[log_dict['阀门井总结']+row].value is not None and  '遗失' in sheet[log_dict['阀门井总结']+row].value:
            check1=1
        if sheet[log_dict['管道防护带总结']+row].value is not None and '深根植物' in sheet[log_dict['管道防护带总结']+row].value:
            check2=1
    temp_list:list[str]=['对示踪系统完整性缺失的问题，应采取相应措施确保管道位置数据的准确性。',]
    if check1 >0:
        temp_list.append('对遗失的阀井进行核实，核实后进行恢复或重建。')
    if check2 >0:
        temp_list.append('对有深根植物伴行的管段，应加强巡查避免其对管道的损害。')
    rg.copy_and_insert_paragraph(doc,'复制建议',temp_list)
    
def solo_main(report_name:str,workbook:Workbook,word,path:str):

    replacements_dict:dict = {}
    replacements_list:list[tuple] = []
    doc_modle_path = f"{config['file']}"
    try:
        doc = word.Documents.Open(doc_modle_path)
        
        if config['para']:
            print('扩张并写入段落')
            do_input_para(doc,workbook,report_name)
        
        print('生成替换用文本')
        replacements_dict |= make_replacement_index(workbook,report_name)
        replacements_list += make_all_replacement_index(workbook,report_name) 

        print('替换内容')
        do_replace( doc , replacements_dict['文本'],replacements_list )
        
        print('附件表格处理')
        do_add_row_for_all(report_name,doc,workbook)

        print('扩张分项报告表格')
        cross_list = expand_all_tables(workbook, doc, report_name)

        print('替换残余内容')
        do_replace( doc , replacements_dict['文本a'])
    
        print('填写分项报告表格')
        do_replace_in_son_report(doc,replacements_dict)
    
        if config['pip_fig']:
            print('扩张路由图')
            expand_all_figs(workbook, doc, report_name)

        print('编制图片替换索引') 
        sign_dict=make_sign_dig_log(workbook,doc,report_name,path,cross_list)

        print('替换所有图片')
        do_replace_all_pic(doc,sign_dict,path)


        # 移动到文档的末端
        selection = word.Selection
        selection.EndKey(6)  # 6 表示 wdStory，即整个文档

        # 更新文档中的所有域
        doc.Fields.Update()
        
        output_file = f"{config['out_path']}\\{report_name}.docx"
        doc.SaveAs2(output_file, FileFormat=16)  # 16 表示docx 17 表示 PDF
        
        # output_file = f"{config['out_path']}\\{report_name}.pdf"
        # doc.SaveAs2(output_file, FileFormat=17)  
        
        print(f"文档已保存为：{output_file}")

    except Exception as ex:
        traceback.print_exc()
        if doc is not None:
            doc.SaveAs2(f"{config['out_path']}\\error_{report_name}.docx",FileFormat =16)
            print(f"{report_name}发生错误！")
            doc.Saved =True
            raise ex
    finally:
        if doc is not None:
            doc.Close(SaveChanges=False)


if __name__ == '__main__':
    global config
    config=set_config.show_config_dialog()
    app_type = rg.check_office_installation()
    if app_type == None:
        print('未找到合适的应用以打开文档')

    path:str = os.getcwd()
    
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\犀浦\\犀浦_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\新繁\\新繁_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\大丰\\大丰_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\管网840\\管网840_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\郫三司\\郫三司_原始数据.xlsx" )
    workbook:Workbook = openpyxl.load_workbook(f"{config['dir']}\\原始数据.xlsx" )

    print('读取模板文件')
    
    if app_type == "office":
        word = win32.Dispatch("Word.Application")
    elif app_type == "wps":
        word = win32.Dispatch("Kwps.Application")
    word.Visible = False  # 不显示 Word 窗口，加快处理速度
    word.DisplayAlerts = 0  # 关闭警告信息
    # 全局关闭拼写/语法检查
    word.Options.CheckSpellingAsYouType = False   # 关闭实时拼写检查
    word.Options.CheckGrammarAsYouType = False    # 关闭实时语法检查
    word.Options.ContextualSpeller = False        # 关闭上下文拼写检查（Word 2010+）
    #   初始化完成
    sheet=workbook['管道基本信息']
    all_names:list[str]=[]
    log_dict =rg.get_col_in_sheet(sheet)
    for cell in sheet[log_dict['报告编号']]:    # 遍历静态台账里所有编号
        v:str = cell.value
        if v is not None:
            all_names.append(v)
        else:
            break
    
    for report_name in all_names[1:]:
        try:
            solo_main(report_name,workbook,word,path)
        except Exception as e:
            print('有错误发生')
        finally:
            continue
    # report_name = 'DGB2025001CD'
    # solo_main(report_name,workbook,word,path)



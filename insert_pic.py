import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
import src.mypackage.r_generator as rg
from pathlib import Path

source:Path=Path(r"E:\BaiduSyncdisk\成渝特检\模板文件与生成程序\记录、报告生成\钢管\破损点\破损点台账.xlsx")
save_path:Path=Path(r"E:\BaiduSyncdisk\成渝特检\模板文件与生成程序\记录、报告生成\钢管\破损点\test.xlsx")
wb:Workbook = openpyxl.load_workbook(str(source))
ws:Worksheet= wb.worksheets[0]

target_rows:list[str] = []
log_dict = rg.get_col_in_sheet(ws) 
for cell in ws[log_dict['ACVG最大dB值']]:
    if cell.value is not None:
        target_rows.append(str(cell.row))
for row in target_rows[1:]:
    image_names:list[str]=ws[log_dict['现场图片（按需求上传）']+row].value.split(',') if ws[log_dict['现场图片（按需求上传）']+row].value else []
    for ist,image_name in enumerate(image_names):
        cell:Cell = ws[log_dict[f'粘贴图片{ist+1}']+row]
        cell.value='点击查看原图'
        # image_name = Path(ws[log_dict['记录自编号']+row].value)
        image_path = '破损点图片'/Path(image_name)
        cell.hyperlink = str(image_path)
        cell.style = "Hyperlink"         
wb.save(str(save_path))

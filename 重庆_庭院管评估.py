"""
    用来生成PE管定检报告
    默认的文件名为：“PE管定检报告模版.docx”
    
"""
import openpyxl
from openpyxl.workbook import Workbook
import datetime
import win32com.client as win32
import os
import traceback
import math
from src.mypackage import r_generator as rg
from src.mypackage.LOG_DATA import LOG_DICT,RISKY_EVA_C_CQ,RISKY_EVA_S_CQ
from src.mypackage import set_config,interraction_terminal 


"""=========================编辑生成全部用于替换的列表索引文件replacements======================"""

def make_text_for_c1(report_name:str,workbook:Workbook):
    """用于生成每个工程最开始的概况文本"""
    sheet= workbook['管段清单']
    log_dict:dict[str] = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    c1_list:list[str] = []
    key_cols:list[str] = [
                    log_dict['工程名称'],
                    '，起止点坐标',log_dict['起止点坐标'],
                    '，竣工图号',log_dict['竣工图号'],
                    '，竣工日期',log_dict['竣工日期'],
                    '，运行年限',log_dict['实际使用年限'],
                    '。'
                    ]
    i:int=0
    for row in rows:
        i += 1
        text:str = rg.get_text_by_log(sheet,row,key_cols)
        c1_list.append(f"{i}、{text}")

    return c1_list

def expand_all_tables(workbook:Workbook, doc, report_name:str)->list[int]:
    """按照读取到的分项报告数量，复制报告页张数。返回穿跨越的组织数量列表"""
    #   宏观检查报告
    sheet = workbook['宏观检查记录']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    times:int = len(rg.get_rows_in_sheet(report_name, sheet ,log_dict['报告编号'])) 
    if times>1:
        rg.copy_and_insert_report(doc , '复制宏观检查报告', times)
    rg.replace_text(doc, '复制宏观检查报告','',2)

    #   开挖检测
    sheet = workbook['开挖检测记录']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    times:int = len(rg.get_rows_in_sheet(report_name , sheet ,log_dict['报告编号']))
    if times>1:
        rg.copy_and_insert_report(doc , '复制开挖报告', times)
    rg.replace_text(doc, '复制开挖报告','',2)  
       
    #   穿跨越检查
    sheet = workbook['宏观检查记录']
    log_dict:dict[str,str] = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    cross_list:list[int]=[] 
    times:int = 0
    for row in rows:
        count:str|None =sheet[log_dict['穿越总结']+row].value
        if count:
            ctrl_value:int =max(math.ceil(count.count('穿越')/9),math.ceil(count.count('跨越')/5))
            times += ctrl_value
            cross_list.append(ctrl_value)
        else:
            cross_list.append(0)
    if times>1:
        rg.copy_and_insert_report(doc , '复制穿跨越报告', times)
    rg.replace_text(doc, '复制穿跨越报告','',2)  
    return cross_list
    
    #   整理删除页面
    # rg.delete_page_by_text(doc, '待删除')





"""
========================执行替换=======================
"""
#   替换文本
def do_replace(doc , replacements1:list[tuple[str,str]],replacements2:list[tuple[str,str]]=[])->None:
    """替换所有文本，先替换全局，再替换单次"""
    for target_text, replacement_text in replacements2:
        rg.replace_text(doc, target_text, replacement_text,2 )
    for target_text, replacement_text in replacements1:
        rg.replace_text(doc, target_text, replacement_text )
    
#   完成索引
def make_replacement_index(workbook:Workbook,report_name:str)->dict:
    """全体替换内容，主要调整函数"""
    replacements:dict[str,list]={}
    replacements['文本'] = []
    #   封面
    temp_list:list[tuple]= []
    sheet =workbook['资料审查']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['序号'])
    row = rows[0]
    global_name = sheet[log_dict['评估单元']+row].value
    #   封面略去，在全局替换中处理    
    
    #   正文
    temp_text:str= f"{global_name}庭院管道，投用于{sheet[log_dict['投用日期']+row].value}年，长度共计约{sheet[log_dict['管道长度']+row].value}。"
    replacements['文本']+=[('+项目概况',temp_text)]*2
    if sheet[log_dict['原始资料及记录审查问题记载']+row].value=='/':
        temp_text=''
    else:
        temp_text=sheet[log_dict['原始资料及记录审查问题记载']+row].value+'，'
    p_z:str = temp_text +f"一年内泄漏抢险{sheet[log_dict['抢险次数']+row].value}次，" +sheet[log_dict['历史检查问题记载']+row].value
    replacements['文本']+=[('+写入资料审查问题',p_z)]

    sheet =workbook['现场检查']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['序号'])
    row = rows[0]
    p_x:str=''
    for i in range(4):
        v=sheet[log_dict[f"问题描述{i+1}"]+row].value
        if v is None or v=='':
            pass
        else:
            p_x+=v+'，'
    p_x=f"{rg.check_text(p_x)}。"
    replacements['文本']+=[('+写入现场检查问题',p_x)]
    
    #   评估结论表
    sheet =workbook['资料审查']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['序号'])
    row = rows[0]
    replacements['老化评估结论表']=[]
    temp_list:list[tuple[str,str|datetime.datetime]]=[]
    temp_list += [('对象简述',f"{global_name}庭院管道，投用于{sheet[log_dict['投用日期']+row].value}年，长度共计约{sheet[log_dict['管道长度']+row].value}。")]
    # sheet = workbook['资料']
    # log_dict = rg.get_col_in_sheet(sheet)
    # rows=rg.get_rows_in_sheet(report_name,sheet,log_dict['编号'])
    # row = rows[0]
    temp_list += [('长度m',f"{sheet[log_dict['管道长度']+row].value}")]
    temp_list += [('管材类别',f"{sheet[log_dict['管道材质']+row].value}")]
    temp_list += [('使用单位',f"{global_name}业主")]
    temp_list += [('$符合安全运行要求',"□符合安全运行要求")]
    temp_list += [('$限期改造','□限期改造')]

    if '立即改造' in f"{sheet[log_dict['评估结论']+row].value}":
        temp_list += [('$落实安全管控措施，可继续运行','□落实安全管控措施，可继续运行')]
        temp_list += [('$立即改造','☑立即改造')]
    else:
        temp_list += [('$立即改造','□立即改造')]
        temp_list += [('$落实安全管控措施，可继续运行','☑落实安全管控措施，可继续运行')]

    temp_list += [('$材质落后','□材质落后')]
    temp_list += [('$使用年限较长','☑使用年限较长')]
    temp_list += [('$腐蚀泄漏严重','☑腐蚀泄漏严重')]
    temp_list += [('$防腐状况较差','☑防腐状况较差')]
    sheet =workbook['现场检查']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['序号'])
    row = rows[0]
    if '占压'  in sheet[log_dict['占压、圈围']+row].value:
        temp_list += [('$建构筑物占压','☑建构筑物占压')]
    else:
        temp_list += [('$建构筑物占压','□建构筑物占压')]
    temp_list += [('$处于/临近地质灾害易发区域','□处于/临近地质灾害易发区域')]
    temp_list += [('$处于/临近人员密集区','☑处于/临近人员密集区')]
    temp_list += [('$其他主要问题：','□其他主要问题：/')]
    replacements['老化评估结论表'].append(temp_list)

    #   资料审查报告 
    replacements['资料审查报告']=[]
    temp_list=[]
    sheet =workbook['资料审查']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['序号'])
    row=rows[0]
    for key in LOG_DICT['重庆评估-资料审查报告填表']:
        temp_list+=[(key,sheet[log_dict[key]+row].value)]
    for key,logs in LOG_DICT['重庆评估-资料审查报告勾选'].items():
        value = sheet[log_dict[key]+row].value
        temp_text = ''
        for log in logs:
            if value in log:
                temp_text += f"☑{log}、"
            else:
                temp_text += f"□{log}、"
        temp_text=rg.check_text(temp_text)
        temp_list += [(key,temp_text)]
    temp_list+=[('历史检查问题记载',f"三年内泄漏抢险{sheet[log_dict['抢险次数']+row].value}次，" +sheet[log_dict['历史检查问题记载']+row].value)]
    replacements['资料审查报告'].append(temp_list) 

    
    #   现场检查报告
    replacements['现场检查报告']=[]
    temp_list=[]
    sheet =workbook['现场检查']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['序号'])
    row=rows[0]
    for key in LOG_DICT['重庆评估-现场检查报告填表']:
        temp_list+=[(key,sheet[log_dict[key]+row].value)]
    temp_list+=[('管道埋深',f"{sheet[log_dict['管道埋深1']+row].value}-{sheet[log_dict['管道埋深2']+row].value}m")]
    for key,logs in LOG_DICT['重庆评估-现场检查报告勾选'].items():
        value = sheet[log_dict[key]+row].value
        temp_text = ''
        for log in logs:
            if log in value:
                temp_text += f"☑{log}、"
            else:
                temp_text += f"□{log}、"
        temp_text=rg.check_text(temp_text)
        temp_list += [(key,temp_text)]
    
    temp_list+=[('问题与结论',p_x)]

    replacements['现场检查报告'].append(temp_list) 
    
    # 风险评估
    sheet =workbook['失效可能性评分']
    log_dict=rg.get_col_in_sheet(sheet)
    rows =rg.get_rows_in_sheet(report_name,sheet,log_dict['序号'])
    replacements['风险评估报告']=[]
    replacements['失效可能性评分']=[]
    temp_list:list[tuple[str,int|float]] = []
    row = rows[0]
    for key_str,list_dict in RISKY_EVA_S_CQ.items():
        risk_score:int=0 
        for son_key,son_tuple in list_dict.items():
            v= sheet[log_dict[son_key]+row].value # 表格实际内容
            for option,score in son_tuple:
                if isinstance(option,tuple): # 如果键是区间（元组）
                    if v>=option[0] and v<option[1]:
                        risk_score += score
                        temp_list += [(son_key,score)] 
                else:
                    if v == option:
                        risk_score += score
                        temp_list += [(son_key,score)]        
        replacements['风险评估报告'].append(risk_score)
    replacements['失效可能性评分'].append(temp_list)
    
    sheet =workbook['失效后果评分']
    log_dict=rg.get_col_in_sheet(sheet)
    rows =rg.get_rows_in_sheet(report_name,sheet,log_dict['序号'])
    row = rows[0]
    replacements['失效后果评分']=[]
    temp_list:list[tuple[str,int|float]] = []
    for key_str,list_dict in RISKY_EVA_C_CQ.items():
        risk_score:float = 0 
        v= sheet[log_dict[key_str]+row].value # 表格实际内容
        for any_tuple in list_dict:
            option,score =any_tuple
            if isinstance(option,tuple): # 如果键是区间（元组）
                if v>=option[0] and v<option[1]:
                    risk_score += score
                    temp_list += [(key_str,score)]   
            else:
                if v == option:
                    risk_score += score 
                    temp_list += [(key_str,score)]   
        replacements['风险评估报告'].append(risk_score)
    replacements['失效后果评分'].append(temp_list)
    s_sigma_value = sum(replacements['风险评估报告'][:5])
    c_sigma_value = sum(replacements['风险评估报告'][5:])
    r_value = s_sigma_value*c_sigma_value
    if r_value<=1600:
        r_class='中低风险'
    elif r_value>1600 and r_value<=3500:
        r_class='中风险'
    elif r_value>3500 and r_value<=7500:
        r_class='中高风险'
    else:
        r_class='高风险'

    replacements['文本']+=[
        ('+评估失效可能性得分',f"{s_sigma_value:.2f}"),
        ('+评估失效后果得分',f"{c_sigma_value:.2f}"),
        ('+评估风险值',f"{r_value:.2f}"),
        ('+评估风险等级',r_class),
        ]*2
    
   
    return replacements

def make_all_replacement_index(workbook,report_name):
    """管道基本信息：报告编号、管道名称、管道长度等"""
    replacements:list = []
    sheet = workbook['资料审查']
    log_dict:dict =rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['序号'])
    row = rows[0]
    replacements += [
                # ('+报告编号',report_name),
                ('+使用单位','重庆江津天然气有限责任公司'),
                ('+投用日期',sheet[log_dict['投用日期']+row].value),
                ('+管道名称',sheet[log_dict['管道名称']+row].value),
                ('+管道长度',sheet[log_dict['管道长度']+row].value ),
                ('+管道规格',sheet[log_dict['管道规格']+row].value ),
                ('+评估结论',f"{sheet[log_dict['评估结论']+row].value}。"),
                # ('+管道长度','164.12'),# 新繁
                # ('+管道名称','天然气管道'),
                # ('+检验日期','2025年06月15日') # 新繁
                # ('+使用单位','成都成燃新繁燃气有限公司'),
                 ]
    sheet = workbook['现场检查']
    log_dict:dict =rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['序号'])
    row = rows[0]
    date:datetime.datetime=sheet[log_dict['检测日期']+row].value
    replacements += [ ('+评估时间',date.strftime("%Y年%m月")),]
    return replacements

def do_replace_in_son_report(doc,any_dict):
    """执行分项报告表格写入"""
    i:int = 0
    j:int = 0
    k:int = 0
    l:int = 1
    m:int = 1
    for table in doc.Tables:
        title_name:str = table.Title
        if title_name == '资料审查报告':
            rg.replace_text_in_table(doc,table,any_dict['资料审查报告'][i],'重庆评估-资料审查报告索引')  
            i+=1
        elif title_name == '现场检查报告':
            rg.replace_text_in_table(doc,table,any_dict['现场检查报告'][j],'重庆评估-现场检查报告索引')  
            j+=1
        elif title_name == '风险评估报告':
            son_table = table.Cell(3,1).Tables(1)
            for score in any_dict['风险评估报告'][:5]:
                cell = son_table.Cell(2,l)
                cell.Range.Text = score
                l+=1
            son_table = table.Cell(3,1).Tables(2)
            for score in any_dict['风险评估报告'][5:]:
                cell = son_table.Cell(2,l-5)
                cell.Range.Text = score
                l+=1
        elif title_name=='老化评估结论表':
             rg.replace_text_in_table(doc,table,any_dict['老化评估结论表'][0],'老化评估结论表') 
        elif title_name=='失效可能性评分':
            rg.replace_text_in_table(doc,table,any_dict['失效可能性评分'][0],'重庆评估-失效可能性评分')  
        elif title_name=='失效后果评分':
            rg.replace_text_in_table(doc,table,any_dict['失效后果评分'][0],'重庆评估-失效后果评分')  
        else:
            pass

def do_replace_all_pic(doc,pic_dict:dict,path:str):
    """执行所有图片的替换"""
    i:int = 0
    j:int = 0
    k:int = 0
    for shape in doc.InlineShapes:
        tag:str = shape.Title 
        if tag == '签字':
            if config['是否生成签字']:
                if pic_dict['签字'][i]=='空白':
                    pass
                else:
                    rg.replace_pictue(doc,f"{config['签名图片所在']}\\{pic_dict['签字'][i]}.png",shape)
            i+=1
        elif tag == '开挖':
            for ex_name in ['.jpg','.png','.jpeg']:
                # f_path:str = f"{path}\\管网840\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                # f_path:str = f"{path}\\新繁\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                # f_path:str = f"{path}\\大丰\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                # f_path:str = f"{path}\\郫三司\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                f_path:str = f"{config['数据源所在']}\\开挖照片\\{pic_dict['开挖'][j]}{ex_name}"
                if os.path.exists(f_path):
                    rg.replace_pictue(doc,f_path,shape,120)
                    break
            j+=1
        elif tag == '管道总图':
            if config['是否写入管道路由图']:
                rg.replace_pictue(doc,pic_dict['管道总图'][0],shape,580)
            pass
        elif tag == '管道分图':
            if config['是否写入管道路由图']:
                rg.replace_pictue(doc,pic_dict['管道分图'][k],shape,580) 
            k+=1
        else:
            pass

def do_input_para(doc,workbook:Workbook,report_name:str)->None:
    sheet = workbook['管段清单']
    log_dict = rg.get_col_in_sheet(sheet)
    rg.copy_and_insert_paragraph(doc,'复制写入概况',make_text_for_c1(report_name,workbook))
    
    sheet =workbook['宏观检查记录']
    log_dict = rg.get_col_in_sheet(sheet)
    rows = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    check1:int =0   # 遗失阀井
    check2:int =0   # 深根植物 
    for row in rows:
        if sheet[log_dict['阀门井总结']+row].value is not None and  '遗失' in sheet[log_dict['阀门井总结']+row].value:
            check1=1
        if sheet[log_dict['管道防护带总结']+row].value is not None and '深根植物' in sheet[log_dict['管道防护带总结']+row].value:
            check2=1
    temp_list:list[str]=['对示踪系统完整性缺失的问题，应采取相应措施确保管道位置数据的准确性。',]
    if check1 >0:
        temp_list.append('对遗失的阀井进行核实，核实后进行恢复或重建。')
    if check2 >0:
        temp_list.append('对有深根植物伴行的管段，应加强巡查避免其对管道的损害。')
    rg.copy_and_insert_paragraph(doc,'复制建议',temp_list)
    
def solo_main(report_name:str,workbook:Workbook,word,path:str):

    replacements_dict:dict = {}
    replacements_list:list[tuple] = []
    doc_modle_path = f"{config['模板文件']}"
    try:
        doc = word.Documents.Open(doc_modle_path)
              
        print('生成替换用文本')
        replacements_dict |= make_replacement_index(workbook,report_name)
        replacements_list += make_all_replacement_index(workbook,report_name) 

        print('替换内容')
        do_replace( doc , replacements_dict['文本'],replacements_list )
        
        print('填写分项报告表格')
        do_replace_in_son_report(doc,replacements_dict)
    
        # 移动到文档的末端
        selection = word.Selection
        selection.EndKey(6)  # 6 表示 wdStory，即整个文档

        # 更新文档中的所有域
        doc.Fields.Update()
        
        output_file = f"{config['输出文件所在']}\\{report_name}.docx"
        doc.SaveAs2(output_file, FileFormat=16)  # 16 表示docx 17 表示 PDF
        
        # output_file = f"{config['输出文件所在']}\\{report_name}.pdf"
        # doc.SaveAs2(output_file, FileFormat=17)  
        
        print(f"文档已保存为：{output_file}")

    except Exception as ex:
        traceback.print_exc()
        if doc is not None:
            doc.SaveAs2(f"{config['输出文件所在']}\\error_{report_name}.docx",FileFormat =16)
            print(f"{report_name}发生错误！")
            doc.Saved =True
            raise ex
    finally:
        if doc is not None:
            doc.Close(SaveChanges=False)


if __name__ == '__main__':
    config:dict[str,str|bool]={
        '模板文件':'',
        '数据源文件':'',
        # '数据源所在':'',
        # '签名图片所在':'',
        '输出文件所在':'',
        # '是否生成概述段落':False,
        # '是否写入封面':False,
        # '是否写入管道清单':False,
        # '是否写入管道路由图':False,
        # '是否生成签字':False,    
    }
    set_list:list[tuple[int,str,str|bool]]=[
        (2,'模板文件','docx'),
        (2,'数据源文件','xlsx'),
        # (0,'数据源所在',''),
        # (0,'签名图片所在',''),
        (0,'输出文件所在',''),
        # (3,'是否生成概述段落',False),
        # (3,'是否写入封面',False),
        # (3,'是否写入管道清单',False),
        # (3,'是否写入管道路由图',False),
        # (3,'是否生成签字',False),    
    ]
    config=interraction_terminal.set_argumments(set_list)
    app_type = rg.check_office_installation()
    if app_type == None:
        print('未找到合适的应用以打开文档')

    path:str = os.getcwd()
    
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\犀浦\\犀浦_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\新繁\\新繁_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\大丰\\大丰_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\管网840\\管网840_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\郫三司\\郫三司_原始数据.xlsx" )
    workbook:Workbook = openpyxl.load_workbook(f"{config['数据源文件']}" )

    print('读取模板文件')
    
    if app_type == "office":
        word = win32.Dispatch("Word.Application")
    elif app_type == "wps":
        word = win32.Dispatch("Kwps.Application")
    word.Visible = False  # 不显示 Word 窗口，加快处理速度
    word.DisplayAlerts = 0  # 关闭警告信息
    # 全局关闭拼写/语法检查
    word.Options.CheckSpellingAsYouType = False   # 关闭实时拼写检查
    word.Options.CheckGrammarAsYouType = False    # 关闭实时语法检查
    word.Options.ContextualSpeller = False        # 关闭上下文拼写检查（Word 2010+）
    #   初始化完成
    sheet=workbook['资料审查']
    all_names:list[str]=[]
    log_dict =rg.get_col_in_sheet(sheet)
    for cell in sheet[log_dict['序号']]:    # 遍历静态台账里所有编号
        v:str = cell.value
        if v is not None:
            all_names.append(v)
        else:
            break
    
    for report_name in all_names[1:]:
        try:
            solo_main(report_name,workbook,word,path)
        except Exception as e:
            print('有错误发生')
        finally:
            continue



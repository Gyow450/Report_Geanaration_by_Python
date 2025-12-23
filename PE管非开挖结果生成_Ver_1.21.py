"""
    用来生成PE管定检报告
    默认的文件名为：“PE管定检报告模版.docx”
    
"""
import openpyxl
from openpyxl.workbook import Workbook
import datetime
import win32com.client as win32
import os
import traceback
import math
from mypackage import r_generator as rg
from mypackage.LOG_DATA import LOG_DICT,RISKY_EVA_C,RISKY_EVA_S
from mypackage import interraction_terminal 


"""=========================编辑生成全部用于替换的列表索引文件replacements======================"""


def expand_all_tables(workbook:Workbook, doc, report_name:str)->None:
    """按照读取到的分项报告数量，复制报告页张数。返回穿跨越的组织数量列表"""
    #   宏观检查报告
    sheet = workbook['宏观检查记录']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    times:int = len(rg.get_rows_in_sheet(report_name, sheet ,log_dict['报告编号'])) 
    
    if times>1:
        rg.copy_and_insert_report(doc , '复制宏观检查报告', times)
    rg.replace_text(doc, '复制宏观检查报告','',2)

    
  
    
    #   整理删除页面
    # rg.delete_page_by_text(doc, '待删除')


"""
========================执行替换=======================
"""
#   替换文本
def do_replace(doc , replacements1:list[tuple[str,str]],replacements2:list[tuple[str,str]]=[])->None:
    """替换所有文本，先替换全局，再替换单次"""
    for target_text, replacement_text in replacements2:
        rg.replace_text(doc, target_text, replacement_text,2 )
    for target_text, replacement_text in replacements1:
        rg.replace_text(doc, target_text, replacement_text )
    

#   完成索引
def make_replacement_index(workbook:Workbook,report_name:str)->dict:
    """全体替换内容，主要调整函数"""
    replacements:dict={}
    replacements['文本'] = []
    
    #   宏观检查报告 
    replacements['宏观检查报告']=[]
    replacements['文本a']=[]
    temp_list:list[tuple]= []
    sheet =workbook['宏观检查记录']
    log_dict:dict = rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])  
    temp_count:int = 0  #分项报告编号
    white_list:set[str] = {'无','符合','完好','合格','正常','保护设施完好'}
    for row in rows:
        
        temp_count += 1
        replacements['文本a']+=[('报告#',f"报告（{temp_count}）")]
        temp_list = [
                    # ('+管道名称',global_name),
                    ('+管道名称',sheet[log_dict['管道名称']+row].value),
                    # ('+管段',sheet[log_dict['管段（桩号）']+row].value),
                    ('+管段','/'),
                    ('+设备名称型号',sheet[log_dict['设备名称型号']+row].value),
                    ('+设备编号',sheet[log_dict['设备编号']+row].value),
                    # ('+检验日期',sheet[log_dict['检验日期']+row].value),
                    ]
        results = ''
        
        for key in LOG_DICT['宏观检查报告']:        #   具体检查项
            result:str = ''
            options:tuple[str] = LOG_DICT['宏观检查报告'][key]      #   体系里所设定：检查项的各种异常选项
            s_text:str|None = sheet[log_dict[key+'总结']+row].value #   多维表格里的单项总结文本 
            if s_text is None or s_text =='':
                check_list:list = []
            elif ',' in s_text:
                check_list:list[str] = s_text.split(',')     #   单选项的总结，用英文,隔断
            else:
                check_list:list[str] = s_text.split(' ')     #   复选项的总结，用空格隔断
            temp_text:str = ''                          #   承载勾选框的文本
            if key not in sheet[log_dict['检查项目总结']+row].value:    #   报告整体总结中的无此项检查
                if '无此项' in options:                 #   若有无此项可选则选，如无则在白名单中匹配
                    for option in options:
                        if option =='无此项':
                            temp_text += f"☑{option}、"
                        else:
                            temp_text += f"□{option}、"
                else:
                    for option in options:
                        if option in white_list:
                            temp_text += f"☑{option}、"
                        else:
                            temp_text += f"□{option}、"
                temp_list += [(f"+{key}",f"{temp_text}□",'  ')]
            else:
                if set(check_list)-white_list:   #   并非只有白名单值
                    extra_options:set[str] =set(check_list)-set(options)
                    for option in options:
                        if option in check_list and option not in white_list:
                            temp_text += f"☑{option}、"
                        else:
                            temp_text += f"□{option}、"
                    if extra_options and '保护设施完好' not in check_list:                   #   存在额外选项
                        if '全线无标志' in extra_options:
                            temp_text=temp_text.replace('□无标志','☑无标志')
                            temp_list += [(f"+{key}",f"{temp_text}□",'  ')]
                        elif '全线深根植物伴行' in extra_options:
                            temp_text=temp_text.replace('□深根植物','☑深根植物')
                            extra_options.discard('全线深根植物伴行')
                            if extra_options:
                                temp_list += [(f"+{key}",f"{temp_text}☑",'，'.join(extra_options))]
                            else:
                                temp_list += [(f"+{key}",f"{temp_text}□",'  ')]
                        else:
                            temp_list += [(f"+{key}",f"{temp_text}☑",'，'.join(extra_options))]
                    else:
                        temp_list += [(f"+{key}",f"{temp_text}□",'  ')]
                    
                        
                else:                                   #   只有白名单值
                    for option in options:
                        if option in white_list:
                            temp_text += f"☑{option}、"
                        else:
                            temp_text += f"□{option}、"
                    temp_list += [(f"+{key}",f"{temp_text}□",'  ')]
            # 统计问题数量
            if s_text != '' and s_text is not None:
                for prblem in set(check_list):
                    if prblem in white_list or prblem in {'无跨越、穿越段仅路面宏观检验','无跨越、穿越段仅宏观检验','暗渠上方跨越，仅地表宏观检查'}:
                        pass
                    else:
                        if '全线无标志' in prblem or '全线深根植物伴行' in prblem: 
                            r_text = f"{prblem}"
                        else:
                            r_text = f"{check_list.count(prblem)}处{prblem}"
                        result += f"{r_text}，"
            if result !='':
                result=rg.check_text(result)
                results += f"{key}存在问题：{result}；"
        
        if results != '':
            results = f"{rg.check_text(results)}；示踪装置：无示踪线。"
        else:
            # results = '以上项目宏观检查未发现异常。'
            results = '示踪装置：无示踪线。'
        temp_list = temp_list + [('+结论',f"结论：{results}")] 
        replacements['宏观检查报告'].append(temp_list) 
    
    return replacements

def make_all_replacement_index(workbook,report_name):
    """管道基本信息：报告编号、管道名称、管道长度等"""
    replacements:list = []
    sheet = workbook['管道基本信息']
    log_dict:dict =rg.get_col_in_sheet(sheet)
    rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    row = rows[0]
    # lenth:int = sheet[log_dict['大于20年长度']+row].value + sheet[log_dict['小于20年长度']+row].value
    lenth:int = sheet[log_dict['管道长度']+row].value 
    l:float = lenth/1000
    replacements += [
                ('+报告编号',report_name),
                ('+使用单位',sheet[log_dict['使用单位']+row].value),
                # ('+使用单位','成都燃气集团股份有限公司管网分公司'),
                ('+检验日期',sheet[log_dict['检验日期']+row].value),
                ('+管道名称',sheet[log_dict['管道名称']+row].value),
                ('+管道长度',l),
                # ('+管道长度','164.12'),# 新繁
                # ('+管道名称','天然气管道'),
                # ('+检验日期','2025年06月15日') # 新繁
                # ('+使用单位','成都成燃新繁燃气有限公司'),
                 ]
    # 检查是否有不明管段
    # sheet = workbook['管段清单']
    # log_dict:dict =rg.get_col_in_sheet(sheet)
    # rows:list[str] = rg.get_rows_in_sheet(report_name,sheet,log_dict['报告编号'])
    # count_unkown:int =0
    # for row in rows:
    #     if '使用单位指定管段' in sheet[log_dict['工程名称']+row].value:
    #         count_unkown =1
    #         break
    # if count_unkown==1:
    #     replacements+=[('+不明管道检查','部分管段无资料，其余管段仅见竣工图')]
    # else:
    #     replacements+=[('+不明管道检查','所有管段仅见竣工图')]
    # replacements+=[('+不明管道检查','所有管段仅见竣工图')]
    # used_years:list[float] = []
    # for row in rows:
    #     if sheet[log_dict['实际使用年限']+row].value:
    #         used_years += [sheet[log_dict['实际使用年限']+row].value]
    # replacements+=[('+投运年限',f"{min(used_years)}—{max(used_years)}年")]
    return replacements

def do_replace_in_son_report(doc,any_dict):
    """执行分项报告表格写入"""
    i:int = 0
    j:int = 0
    k:int = 0
    l:int = 1
    m:int = 1
    for table in doc.Tables:
        title_name:str = table.Title
        if title_name == '宏观检查报告':
            rg.replace_text_in_table(doc,table,any_dict['宏观检查报告'][i],'宏观检查报告索引')  
            i+=1
        elif title_name == '开挖检验报告':
            rg.replace_text_in_table(doc,table,any_dict['开挖检验报告'][j],'开挖检测报告索引')  
            j+=1
        elif title_name == '穿、跨越报告':
            rg.replace_text_in_table(doc,table,any_dict['穿、跨越报告'][k],'穿、跨越报告索引')  
            k+=1
        elif title_name == '风险预评估报告':
            son_table = table.Cell(3,1).Tables(1)
            for score in any_dict['风险预评估'][:8]:
                cell = son_table.Cell(2,l)
                cell.Range.Text = score
                l+=1
            son_table = table.Cell(3,1).Tables(2)
            for score in any_dict['风险预评估'][8:]:
                cell = son_table.Cell(2,l-8)
                cell.Range.Text = score
                l+=1
        elif title_name == '风险再评估报告':
            son_table = table.Cell(3,1).Tables(1)
            for score in any_dict['风险再评估'][:8]:
                cell = son_table.Cell(2,m)
                cell.Range.Text = score
                m+=1
            son_table = table.Cell(3,1).Tables(2)
            for score in any_dict['风险再评估'][8:]:
                cell = son_table.Cell(2,m-8)
                cell.Range.Text = score
                m+=1
        else:
            pass




    
def solo_main(report_name:str,workbook:Workbook,word,path:str):

    replacements_dict:dict = {}
    replacements_list:list[tuple] = []
    doc_modle_path = f"{config['模板文件']}"
    try:
        doc = word.Documents.Open(doc_modle_path)
        
     
        print('生成替换用文本')
        replacements_dict |= make_replacement_index(workbook,report_name)
        # replacements_list += make_all_replacement_index(workbook,report_name) 

        print('替换内容')
        do_replace( doc , replacements_dict['文本'],replacements_list )
        

        print('扩张分项报告表格')
        expand_all_tables(workbook, doc, report_name)

        print('替换残余内容')
        do_replace( doc , replacements_dict['文本a'])
    
        print('填写分项报告表格')
        do_replace_in_son_report(doc,replacements_dict)


        # 移动到文档的末端
        selection = word.Selection
        selection.EndKey(6)  # 6 表示 wdStory，即整个文档

        # 更新文档中的所有域
        doc.Fields.Update()
        
        output_file = f"{config['输出文件所在']}\\{report_name}.docx"
        doc.SaveAs2(output_file, FileFormat=16)  # 16 表示docx 17 表示 PDF
        
        # output_file = f"{config['输出文件所在']}\\{report_name}.pdf"
        # doc.SaveAs2(output_file, FileFormat=17)  
        
        print(f"文档已保存为：{output_file}")

    except Exception as ex:
        traceback.print_exc()
        if doc is not None:
            doc.SaveAs2(f"{config['输出文件所在']}\\error_{report_name}.docx",FileFormat =16)
            print(f"{report_name}发生错误！")
            doc.Saved =True
            raise ex
    finally:
        if doc is not None:
            doc.Close(SaveChanges=False)


if __name__ == '__main__':
    config:dict[str,str|bool]={
        '模板文件':'',
        '数据源文件':'',
        # '签名图片所在':'',
        '输出文件所在':'',
        # '是否生成概述段落':False,
        # '是否写入封面':False,
        # '是否写入管道清单':False,
        # '是否写入管道路由图':False,
        # '是否生成签字':False,    
    }
    set_list:list[tuple[int,str,str|bool]]=[
        (2,'模板文件','docx'),
        (2,'数据源文件','xlsx'),
        # (0,'签名图片所在',''),
        (0,'输出文件所在',''),
        # (3,'是否生成概述段落',False),
        # (3,'是否写入封面',False),
        # (3,'是否写入管道清单',False),
        # (3,'是否写入管道路由图',False),
        # (3,'是否生成签字',False),    
    ]
    config=interraction_terminal.set_argumments(set_list)
    app_type = rg.check_office_installation()
    if app_type == None:
        print('未找到合适的应用以打开文档')

    path:str = os.getcwd()
    
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\犀浦\\犀浦_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\新繁\\新繁_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\大丰\\大丰_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\管网840\\管网840_原始数据.xlsx" )
    # workbook:Workbook = openpyxl.load_workbook(f"{path}\\郫三司\\郫三司_原始数据.xlsx" )
    workbook:Workbook = openpyxl.load_workbook(f"{config['数据源文件']}" )

    print('读取模板文件')
    
    if app_type == "office":
        word = win32.Dispatch("Word.Application")
    elif app_type == "wps":
        word = win32.Dispatch("Kwps.Application")
    word.Visible = False  # 不显示 Word 窗口，加快处理速度
    word.DisplayAlerts = 0  # 关闭警告信息
    # 全局关闭拼写/语法检查
    word.Options.CheckSpellingAsYouType = False   # 关闭实时拼写检查
    word.Options.CheckGrammarAsYouType = False    # 关闭实时语法检查
    word.Options.ContextualSpeller = False        # 关闭上下文拼写检查（Word 2010+）
    #   初始化完成
    sheet=workbook['管道基本信息']
    all_names:list[str]=[]
    log_dict =rg.get_col_in_sheet(sheet)
    for cell in sheet[log_dict['报告编号']]:    # 遍历静态台账里所有编号
        v:str = cell.value
        if v is not None:
            all_names.append(v)
        else:
            break
    
    for report_name in set(all_names[1:]):
        try:
            solo_main(report_name,workbook,word,path)
        except Exception as e:
            print('有错误发生')
        finally:
            continue
    # report_name = 'DGB2025001CD'
    # solo_main(report_name,workbook,word,path)


